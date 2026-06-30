import streamlit as st
import io
from io import BytesIO
import tempfile
import zipfile
import os
import json
import pyreadstat
import traceback
import pandas as pd
import re

from datetime import datetime


from schema_builder import build_rawdata_spec

from pipeline.stat_builder import (
    process_raw_data
)

from pipeline.datacheck_builder import (
    parse_dvp_file
)

from pipeline.datacheck_builder import (
    parse_dvp_file,
    build_dvp_review_df,
    run_dvp_data_check,
    build_unified_fail_report_df,
)


from verify.verify import (
    compare_excel_verify,
    export_verify_report,
    build_summary_view,
    get_structure_diff_summary
)



# =========================================================
# Helper functions
# =========================================================
# 抓檔名 (Sponsor, Protocol)
def extract_protocol_no_from_filename(file_name):
    if not file_name:
        return ""

    name = os.path.splitext(
        file_name
    )[0].strip()

    parts = [
        p.strip()
        for p in name.split("_")
        if p.strip()
    ]

    if not parts:
        return "", ""


    # --------------------------------------
    # sponsor_protocol_xxx
    # --------------------------------------
    if len(parts) >= 4:

        sponsor = parts[0]
        protocol = parts[1]

    # --------------------------------------
    # protocol_xxx
    # --------------------------------------
    elif len(parts) >= 2:

        sponsor = ""
        protocol = parts[0]

    else:

        sponsor = ""
        protocol = parts[0]

    return sponsor, protocol






def Export_excel(sheet_dict):
    output = BytesIO()

    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheet_dict.items():

            # 如果是空 df 就 skip（避免空 sheet）
            if df is None or df.empty:
                continue

            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.book[sheet_name]

            # Freeze first row
            ws.freeze_panes = "A2"

            # Auto filter（非常推薦）
            ws.auto_filter.ref = ws.dimensions

            # Header style
            header_fill = PatternFill(
                start_color="F4A300",
                end_color="F4A300",
                fill_type="solid"
            )

            header_font = Font(name="Calibri", size=10, bold=True)
            normal_font = Font(name="Calibri", size=10)

            align_wrap_top = Alignment(wrap_text=True, vertical="top")
            align_wrap_center = Alignment(wrap_text=True, vertical="center")

            # Header style
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_wrap_center

            # Body style（避免重複設定 header）
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = normal_font
                    cell.alignment = align_wrap_top

            # 欄寬自動（稍微 smarter）
            for col_idx, col in enumerate(ws.columns, start=1):
                max_length = 0

                col_letter = get_column_letter(col_idx)

                for cell in col:
                    try:
                        if cell.value:
                            val = str(cell.value)
                            max_length = max(max_length, len(val))
                    except:
                        pass

                # 限制最大寬度（避免爆炸）
                ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

            # Row height（改善 wrap）
            for row in ws.iter_rows():
                max_lines = 1

                for cell in row:
                    if cell.value:
                        text = str(cell.value)
                        approx_lines = max(
                            text.count("\n") + 1,
                            (len(text) // 40) + 1
                        )
                        max_lines = max(max_lines, approx_lines)

                ws.row_dimensions[row[0].row].height = max_lines * 15

    output.seek(0)
    return output.getvalue()
    # End=========================================================






def style_compare(df: pd.DataFrame):
    """
    highlight compare table
    """
    def highlight_row(row):
        status = str(row.get("STATUS", "")).upper()

        if status == "MISSING_IN_RAW":
            return ["background-color: #f8d7da"] * len(row)   # 淡紅
        elif status == "EXTRA_IN_RAW":
            return ["background-color: #fff3cd"] * len(row)   # 淡黃
        elif status == "DATATYPE_MISMATCH":
            return ["background-color: #d1ecf1"] * len(row)   # 淡藍
        elif status == "MATCH":
            return ["background-color: #d4edda"] * len(row)   # 淡綠
        elif status == "SYSTEM_VAR":
            return ["background-color: #e2e3e5"] * len(row)   # 淡灰
        else:
            return [""] * len(row)

    return df.style.apply(highlight_row, axis=1)


def apply_filters(df, dataset_filter, status_filter, keyword, only_issues):
    df = df.copy()

    if dataset_filter:
        df = df[df["DATASET"].isin(dataset_filter)]

    if status_filter:
        df = df[df["STATUS"].isin(status_filter)]
   
    if only_issues:
        df = df[~df["STATUS"].isin(["MATCH", "SYSTEM_VAR"])]

    if keyword:
        k = keyword.upper()
        df = df[
            df["DATASET"].astype(str).str.contains(k, na=False) |
            df["VARIABLE"].astype(str).str.contains(k, na=False) |
            df["LABEL"].astype(str).str.contains(k, na=False)
        ]

    return df



def render_summary(compare_df):

    total = len(compare_df)

    match_cnt = (compare_df["STATUS"] == "MATCH").sum()
    mismatch_cnt = (compare_df["STATUS"] == "DATATYPE_MISMATCH").sum()
    missing_cnt = (compare_df["STATUS"] == "MISSING_IN_RAW").sum()
    extra_cnt = (compare_df["STATUS"] == "EXTRA_IN_RAW").sum()
    system_cnt = int((compare_df["STATUS"] == "SYSTEM_VAR").sum())

    c1, c2, c3, c4, c5, c6 = st.columns(6)

    c1.metric("Total", total)
    c2.metric("🟩 Match", match_cnt)
    c3.metric("🟥 Missing in Raw", missing_cnt)
    c4.metric("🟨 Extra in Raw", extra_cnt)
    c5.metric("🟦 Data Type Mismatch", mismatch_cnt)
    c6.metric("⬜ System", system_cnt)



def reorder_preview_columns(df):
    priority = ["USUBJID", "SUBJID", "SOURCEDN", "VISIT", "VISITNUM", "SITEID", "STUDYID"]
    exist = [c for c in priority if c in df.columns]
    others = [c for c in df.columns if c not in exist]
    return df[exist + others]



def build_merged_excel_bytes(merged_dataset_map: dict) -> bytes:
    """
    所有 merged datasets 輸出到同一個 Excel，
    每個 dataset 一個 sheet。
    """
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, df in merged_dataset_map.items():
            if df is None:
                continue

            # Excel sheet 名稱限制 31 字元，且不能有某些特殊字元
            sheet_name = str(name).strip()
            for ch in [":", "\\", "/", "?", "*", "[", "]"]:
                sheet_name = sheet_name.replace(ch, "_")
            sheet_name = sheet_name[:31] if sheet_name else "DATASET"

            df.to_excel(writer, sheet_name=sheet_name, index=False)

    buffer.seek(0)
    return buffer.getvalue()


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    df2 = df.copy()

    for col in df2.columns:
        s = df2[col]

        # float 保留小數（避免 1.0 → 1）
        if "float" in str(s.dtype):
            df2[col] = s.map(lambda x: "" if pd.isna(x) else format(x, ".15g"))

        # datetime → ISO（SAS最好讀）
        elif "datetime" in str(s.dtype):
            df2[col] = s.dt.strftime("%Y-%m-%dT%H:%M:%S")

    return df2.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")




def sanitize_sas_dataset_name(name: str, max_len=32) -> str:
    s = str(name).strip().upper()

    for ch in [" ", "-", "/", "\\", "(", ")", "[", "]", "{", "}", ".", ",", ":", ";", "?", "*", "+", "&"]:
        s = s.replace(ch, "")

    s = "".join(ch for ch in s if (ch.isalnum() or ch == "_"))

    if not s:
        s = "DATASET"

    if not s[0].isalpha():
        s = "X" + s

    return s[:max_len]


def infer_char_length(series: pd.Series) -> int:
    if series is None or len(series) == 0:
        return 1

    non_null = series.dropna()
    if len(non_null) == 0:
        return 1

    max_len = non_null.astype(str).map(len).max()

    if pd.isna(max_len):
        return 1

    return max(1, int(max_len))


def infer_sas_type_from_raw(raw_list: list) -> str:
    """
    用 raw_df 的 RAW_DATATYPE_STD 判斷
    """
    raw_list = [str(x).upper() for x in raw_list]

    if any(x in ["TEXT", "DATE", "DATETIME", "TIME"] for x in raw_list):
        return "char"

    if any(x == "NUMERIC" for x in raw_list):
        return "num"

    return "char"


def get_source_datasets(dataset_name, merge_log_df):
    """
    找 merged dataset 對應來源 dataset
    """
    if merge_log_df is None or merge_log_df.empty:
        return [dataset_name.upper()]

    tmp = merge_log_df[
        merge_log_df["RESULT_DATASET"].astype(str).str.upper()
        == dataset_name.upper()
    ]

    if tmp.empty:
        return [dataset_name.upper()]

    sub = str(tmp.iloc[0]["SUBITEMS"]).split(",")

    return [x.strip().upper() for x in sub if x.strip()]




def prepare_view_df(df, merge_mode):
    if df is None or df.empty:
        return df

    view_df = reorder_preview_columns(df)

    if merge_mode == "horizontal" and "SOURCEDN" in view_df.columns:
        view_df = view_df.drop(columns=["SOURCEDN"])

    return view_df


# =========================================================
# 單一 dataset metadata
# =========================================================
def build_dataset_metadata(dataset_name, df, raw_df=None, merge_log_df=None):

    dataset_name = str(dataset_name).strip()
    sas_name = sanitize_sas_dataset_name(dataset_name)

    source_datasets = get_source_datasets(dataset_name, merge_log_df)

    ds_meta = {
        "dataset": dataset_name,
        "sas_dataset": sas_name,
        "csv_file": f"{dataset_name}.csv",
        "nobs": int(len(df)),
        "nvars": int(len(df.columns)),
        "variables": []
    }

    for col in df.columns:

        col_upper = str(col).strip().upper()
        s = df[col]

        # =================================================
        # 新增 derived variables：直接指定型態
        # =================================================
        if col_upper == "VISIT_DVP":
            sas_type = "char"
            var_meta = {
                "name": col_upper,
                "sas_type": sas_type,
                "source": "derived",
                "derivation": "visit mapping"
            }

        elif col_upper == "VISITNUM":
            sas_type = "num"
            var_meta = {
                "name": col_upper,
                "sas_type": sas_type,
                "source": "derived",
                "derivation": "infer_visitnum_from_raw_visit"
            }

        # --- DTC char ---
        elif col_upper.endswith("DTC_C"):
            sas_type = "char"
            var_meta = {
                "name": col_upper,
                "sas_type": sas_type,
                "source": "derived",
                "derivation": "ISO8601 datetime (char)"
            }

        # --- DTC numeric ---
        elif col_upper.endswith("DTC_N"):
            sas_type = "num"
            var_meta = {
                "name": col_upper,
                "sas_type": sas_type,
                "source": "derived",
                "derivation": "ISO8601 datetime (SAS datetime)"
            }
            var_meta["format"] = "E8601DT19."
            ds_meta["variables"].append(var_meta)
            continue

        # --- TIM char ---
        elif col_upper.endswith("TIM_C"):
            sas_type = "char"
            var_meta = {
                "name": col_upper,
                "sas_type": sas_type,
                "source": "derived",
                "derivation": "time normalized (char)"
            }

        # --- TIM numeric ---
        elif col_upper.endswith("TIM_N"):
            sas_type = "num"
            var_meta = {
                "name": col_upper,
                "sas_type": sas_type,
                "source": "derived",
                "derivation": "time normalized (SAS time)"
            }
            var_meta["format"] = "TIME8."
            ds_meta["variables"].append(var_meta)
            continue

        # --- 一般 char date ---
        elif col_upper.endswith("_C"):
            sas_type = "char"
            var_meta = {
                "name": col_upper,
                "sas_type": sas_type,
                "source": "derived",
                "derivation": "partial date normalized (char)"
            }

        # --- 一般 numeric date ---
        elif col_upper.endswith("_N"):
            sas_type = "num"
            var_meta = {
                "name": col_upper,
                "sas_type": sas_type,
                "source": "derived",
                "derivation": "partial date normalized (SAS date)"
            }
            var_meta["format"] = "YYMMDD10."
            ds_meta["variables"].append(var_meta)
            continue


        # =================================================
        # 其他欄位：照原本 raw_df 推斷
        # =================================================
        else:
            sas_type = None

            # 優先從 raw_df 找原始型態
            if raw_df is not None and not raw_df.empty:

                tmp = raw_df[
                    raw_df["VARIABLE"].astype(str).str.upper() == col_upper
                ]

                if "DATASET" in raw_df.columns:
                    tmp = tmp[
                        tmp["DATASET"].astype(str).str.upper().isin(source_datasets)
                    ]

                if not tmp.empty and "RAW_DATATYPE_STD" in tmp.columns:
                    raw_types = tmp["RAW_DATATYPE_STD"].unique().tolist()
                    sas_type = infer_sas_type_from_raw(raw_types)

            # fallback
            if sas_type is None:
                if pd.api.types.is_numeric_dtype(s) and not pd.api.types.is_bool_dtype(s):
                    sas_type = "num"
                else:
                    sas_type = "char"

            var_meta = {
                "name": col_upper,
                "sas_type": sas_type
            }

        if sas_type == "char":
            length = 2000
            var_meta["length"] = length
            var_meta["informat"] = f"$CHAR{length}."
            var_meta["format"] = f"${length}."
        else:
            var_meta["informat"] = "BEST32."
            var_meta["format"] = "BEST32."

        ds_meta["variables"].append(var_meta)

    return ds_meta


# =========================================================
# 外部用：產 metadata.json text
# =========================================================
def build_metadata_json(merged_dataset_map, raw_df=None, merge_log_df=None):

    out = {"version": 1, "datasets": []}

    for name, df in merged_dataset_map.items():

        if df is None:
            continue

        ds_meta = build_dataset_metadata(
            name, df, raw_df=raw_df, merge_log_df=merge_log_df
        )

        out["datasets"].append(ds_meta)

    return json.dumps(out, indent=2, ensure_ascii=False)





def build_sas_script():

    return r"""
%macro grabpath;
  %qsubstr(%sysget(SAS_EXECFILEPATH), 1,
    %length(%sysget(SAS_EXECFILEPATH))
    - %length(%sysget(SAS_EXECFILEname))
  )
%mend;

%let path=%grabpath;
%let csvdir=&path;
%let metajson=&path.metadata.json;

/* === JSON read === */
filename METAJSON "&metajson";
libname META JSON fileref=METAJSON;

/* === dataset list === */
proc sql noprint;
  select dataset, sas_dataset, csv_file
    into :DS1-:DS999,
         :SASDS1-:SASDS999,
         :CSV1-:CSV999
  from META.datasets;
  %let NDS=&sqlobs;
quit;


/* =====================================================
   import one dataset
===================================================== */
%macro import_one(ds, sasds, csvfile);

    %put ===== Import &ds -> &sasds =====;

  /* -------------------------------------------------
     1) 讀 header（只讀第一列）
  ------------------------------------------------- */
  filename FIN "&csvdir.&csvfile";

  data _header;
    infile FIN dsd dlm=',' truncover obs=1 lrecl=32767;
    input;
    header = _infile_;
  run;

  data _header_vars;
    set _header;
    length varname $200;
    do ordinal_header = 1 to countw(header, ',');
      varname = dequote(strip(scan(header, ordinal_header, ',')));
      output;
    end;
    keep ordinal_header varname;
  run;

  /* -------------------------------------------------
     2) metadata variables（依 dataset 取出）
        注意：這裡讀的是 META.datasets_variables
  ------------------------------------------------- */
  proc sql;
    create table _varmeta as
    select d.dataset,
           d.sas_dataset,
           v.ordinal_variables,
           upcase(v.name) as meta_name length=64,
           lowcase(v.sas_type) as sas_type length=16,
           coalesce(v.length, 200) as var_length,
           coalesce(v.informat, '') as informat length=32,
           coalesce(v.format,   '') as format   length=32
    from META.datasets d
    inner join META.datasets_variables v
      on d.ordinal_datasets = v.ordinal_datasets
    where upcase(d.dataset) = upcase("&ds")
    order by v.ordinal_variables
    ;
  quit;

  /* -------------------------------------------------
     3) 用 CSV header 順序對應 metadata
        這一步就是避免從 VISIT 開始整列歪掉的關鍵
  ------------------------------------------------- */
  proc sql;
    create table _import_meta as
    select h.ordinal_header,
           upcase(h.varname) as varname length=200,
           coalesce(m.sas_type, 'char') as sas_type length=16,
           coalesce(m.var_length, 200) as var_length,
           coalesce(m.informat, '') as informat length=32,
           coalesce(m.format, '') as format length=32
    from _header_vars h
    left join _varmeta m
      on upcase(h.varname) = m.meta_name
    order by h.ordinal_header
    ;
  quit;

  /* -------------------------------------------------
     4) 產 LENGTH / INPUT macro vars（照 header 順序）
        這裡所有 CHAR 都固定給 200
  ------------------------------------------------- */
  proc sql noprint;
    select varname,
           sas_type,
           case when sas_type='char' then 200 else 8 end,
           informat,
           format
      into :VAR1-:VAR999,
           :TYPE1-:TYPE999,
           :LEN1-:LEN999,
           :INF1-:INF999,
           :FMT1-:FMT999
    from _import_meta
    order by ordinal_header
    ;
    %let NVAR=&sqlobs;
  quit;

  %let LENGTH_STMT=;
  %let INPUT_STMT=;

  %do j=1 %to &NVAR;
    %if %upcase(&&TYPE&j)=CHAR %then %do;
      %let LENGTH_STMT=&LENGTH_STMT %superq(VAR&j) $200;
      %let INPUT_STMT=&INPUT_STMT %superq(VAR&j) :$char200.;
    %end;
    %else %do;
      %let LENGTH_STMT=&LENGTH_STMT %superq(VAR&j) 8;
      %let INPUT_STMT=&INPUT_STMT %superq(VAR&j) :?? best32.;
    %end;
  %end;

  %put NOTE: LENGTH_STMT=&LENGTH_STMT;
  %put NOTE: INPUT_STMT=&INPUT_STMT;

  /* -------------------------------------------------
     5) 先清洗 CSV（保留你原本雙引號處理）
  ------------------------------------------------- */
  filename FILE_TMP temp;

  data _null_;
    infile FIN lrecl=32767 truncover;
    file FILE_TMP lrecl=32767;
    input line $char32767.;
    line = transtrn(line,'""','""');
    put line;
  run;

  /* -------------------------------------------------
     6) 正式匯入（不用 PROC IMPORT）
        完全照 metadata + header 順序
  ------------------------------------------------- */
  data &sasds;
    infile FILE_TMP
      dsd
      dlm=','
      truncover
      firstobs=2
      lrecl=32767
    ;

    length &LENGTH_STMT;
    input &INPUT_STMT;
  run;

  /* -------------------------------------------------
     7) 套 format / informat
  ------------------------------------------------- */
  proc datasets library=work nolist;
    modify &sasds;
    %do j=1 %to &NVAR;
      %if %superq(FMT&j) ne %then %do;
        format &&VAR&j &&FMT&j;
      %end;
      %if %superq(INF&j) ne %then %do;
        informat &&VAR&j &&INF&j;
      %end;
    %end;
  quit;

  filename FIN clear;
  filename FILE_TMP clear;

%mend;


/* =====================================================
   import all datasets
===================================================== */
%macro import_all;
  %do i=1 %to &NDS;
    %let ds=&&DS&i;
    %let sasds=&&SASDS&i;
    %let csvfile=&&CSV&i;

    %put &ds &sasds. &csvfile.;

    %import_one(&ds, &sasds, &csvfile);
  %end;
%mend;

%import_all;

ods exclude Base.Datasets.Directory Base.Datasets.Members;
proc datasets library=work;
  delete _header: _import_meta _varmeta;
quit;
"""






def build_csv_zip_bytes(
    merged_dataset_map,
    raw_df,
    merge_log_df
):
    zip_buffer = io.BytesIO()

    # metadata
    metadata_json_text = build_metadata_json(
        merged_dataset_map=merged_dataset_map,
        raw_df=raw_df,
        merge_log_df=merge_log_df
    )

    # SAS script
    sas_script_text = build_sas_script()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:

        # =========================
        # CSV files
        # =========================
        for name, df in merged_dataset_map.items():

            if df is None:
                continue

            csv_bytes = dataframe_to_csv_bytes(df)

            safe_name = str(name).strip()
            zf.writestr(f"{safe_name}.csv", csv_bytes)

        # =========================
        # metadata.json
        # =========================
        zf.writestr(
            "metadata.json",
            metadata_json_text.encode("utf-8")
        )

        # =========================
        # SAS script
        # =========================
        zf.writestr(
            "RunMe.sas",
            sas_script_text.encode("utf-8")
        )

    zip_buffer.seek(0)
    return zip_buffer.getvalue()




def refs_to_df(refs):

    if not refs:
        return pd.DataFrame(columns=[
            "RAW_TOKEN",
            "BASE_VAR",
            "REF_STAT_DATASET",
            "JOIN_TYPE",
            "RESOLVED_JOIN_TYPE",
            "RESOLVED_JOIN_KEYS",
            "VISIT_SUFFIX",
            "PARAM"
        ])

    df = pd.DataFrame(refs)

    # 展平 KEY list
    if "RESOLVED_JOIN_KEYS" in df.columns:
        df["RESOLVED_JOIN_KEYS"] = df["RESOLVED_JOIN_KEYS"].apply(
            lambda x: ", ".join(x) if isinstance(x, list) else str(x)
        )

    return df







# =========================================================
# 用於Verify
# =========================================================
def style_summary(df):

    if df.empty:
        return df

    def _to_number(value):
        try:
            if value in ["", None]:
                return 0
            return float(value)
        except Exception:
            return 0

    def _is_non_empty(value):
        return value is not None and str(value).strip() != ""

    def highlight(row):
        status = str(row.get("Status", ""))

        is_inconsistent_row = status != "Consistent"

        base_style = "background-color: #F2F2F2;" if is_inconsistent_row else ""

        styles = [base_style for _ in row.index]

        col_pos = {
            col: idx
            for idx, col in enumerate(row.index)
        }

        def mark_red(col):
            if col in col_pos:
                styles[col_pos[col]] = f"{styles[col_pos[col]]} color: red;"


        # Number of Variable: compare left vs right
        variable_cols = [
            c for c in row.index
            if c.startswith("Number of Variable")
        ]

        if len(variable_cols) == 2:
            left_col, right_col = variable_cols

            if _to_number(row[left_col]) != _to_number(row[right_col]):
                mark_red(left_col)
                mark_red(right_col)

        # Number of Row: compare left vs right
        row_cols = [
            c for c in row.index
            if c.startswith("Number of Row")
        ]

        if len(row_cols) == 2:
            left_col, right_col = row_cols

            if _to_number(row[left_col]) != _to_number(row[right_col]):
                mark_red(left_col)
                mark_red(right_col)


        # Number of Dup. Row: side-specific, mark if > 0
        dup_cols = [
            c for c in row.index
            if c.startswith("Number of Dup. Row")
        ]

        for col in dup_cols:
            if _to_number(row[col]) > 0:
                mark_red(col)


        # Diff Variables: side-specific, mark if not blank
        diff_var_cols = [
            c for c in row.index
            if c.startswith("Diff Variables")
        ]

        for col in diff_var_cols:
            if _is_non_empty(row[col]):
                mark_red(col)

        # Number of Different Rows: mark if > 0
        if "Number of Different Rows" in row.index:
            if _to_number(row["Number of Different Rows"]) > 0:
                mark_red("Number of Different Rows")


        return styles

    return df.style.apply(highlight, axis=1)


def style_data_diff(df, result):
    """
    UI styling for Data Diff.
    """

    if df.empty:
        return df

    labels = result.get("Labels", {})

    left = labels.get("LEFT", "Main")
    right = labels.get("RIGHT", "Double")

    left_u = left.upper()
    right_u = right.upper()

    only_left_label = labels.get("ONLY_LEFT", f"Only in {left}")
    only_right_label = labels.get("ONLY_RIGHT", f"Only in {right}")

    metadata_cols = {
        "DIFF_TYPE",
        "SOURCE",
    }

    value_cols = [
        col for col in df.columns
        if col not in metadata_cols
    ]

    # --------------------------------------------------
    # 建立空 style matrix
    # --------------------------------------------------
    styles = pd.DataFrame(
        "",
        index=df.index,
        columns=df.columns,
    )


    # --------------------------------------------------
    # Row background
    # --------------------------------------------------
    for idx in df.index:

        source = (
            str(df.at[idx, "SOURCE"])
            if "SOURCE" in df.columns
            else ""
        )

        diff_type = (
            str(df.at[idx, "DIFF_TYPE"])
            if "DIFF_TYPE" in df.columns
            else ""
        )

        # Duplicate Record
        if diff_type == "Duplicate Record":

            if source == left:

                styles.loc[idx, :] = (
                    styles.loc[idx, :]
                    + "background-color:#E2F0D9;"
                )

            elif source == right:

                styles.loc[idx, :] = (
                    styles.loc[idx, :]
                    + "background-color:#FCE4D6;"
                )

        # Normal rows
        elif source == right:

            styles.loc[idx, :] = (
                styles.loc[idx, :]
                + "background-color:#F2F2F2;"
            )


    # --------------------------------------------------
    # 分群
    # --------------------------------------------------
    diff_type_series = df.get("DIFF_TYPE", pd.Series("", index=df.index)).astype(str)

    left_only_rows = df.index[
        diff_type_series == only_left_label
    ].tolist()

    right_only_rows = df.index[
        diff_type_series == only_right_label
    ].tolist()

    record_count_diff_rows = df.index[
        diff_type_series == "Record Count Difference"
    ].tolist()

    # --------------------------------------------------
    # Helper
    # --------------------------------------------------
    def _cell_text(row_idx, col):
        value = df.at[row_idx, col]
        return "" if pd.isna(value) else str(value)

    def _similarity_score(row_a, row_b):
        score = 0

        for col in value_cols:
            if _cell_text(row_a, col) == _cell_text(row_b, col):
                score += 1

        return score

    def _mark_red(row_idx, col):
        if col in styles.columns:
            styles.at[row_idx, col] = styles.at[row_idx, col] + "color:red;"

    # --------------------------------------------------
    # Pair Only-in-left vs Only-in-right
    # --------------------------------------------------
    used_right_rows = set()

    for left_row in left_only_rows:

        best_right_row = None
        best_score = -1

        for right_row in right_only_rows:
            if right_row in used_right_rows:
                continue

            score = _similarity_score(left_row, right_row)

            if score > best_score:
                best_score = score
                best_right_row = right_row

        # 找不到 pair：整筆 original data cells 都紅
        if best_right_row is None:
            for col in value_cols:
                _mark_red(left_row, col)
            continue

        used_right_rows.add(best_right_row)

        # 成對後，只標真正不同的 cell
        for col in value_cols:
            left_value = _cell_text(left_row, col)
            right_value = _cell_text(best_right_row, col)

            if left_value != right_value:
                _mark_red(left_row, col)
                _mark_red(best_right_row, col)

    # right-only 沒有被配對：整筆 original data cells 都紅
    for right_row in right_only_rows:
        if right_row not in used_right_rows:
            for col in value_cols:
                _mark_red(right_row, col)

    # --------------------------------------------------
    # Record Count Difference
    # 只標 COUNT / ROWS，不標 original value columns
    # --------------------------------------------------
    count_left_col = f"COUNT_{left_u}"
    count_right_col = f"COUNT_{right_u}"
    rows_left_col = f"ROWS_{left_u}"
    rows_right_col = f"ROWS_{right_u}"

    for row_idx in record_count_diff_rows:
        for col in [
            count_left_col,
            count_right_col,
            rows_left_col,
            rows_right_col,
        ]:
            if col in df.columns:
                _mark_red(row_idx, col)

    return df.style.apply(
        lambda _: styles,
        axis=None,
    )




# =========================================================
# Page config
# =========================================================
st.set_page_config(
    page_title="BS Process",
    layout="wide"
)


# Sidebar Navigation
st.sidebar.title("📊 BS Process")

page = st.sidebar.radio(
    "Select Function",
    [
        "Raw Data SPEC List",
        "Data Process",
        "DVP Review and Data Check",
        "Verify"
    ]
)

today_str = datetime.now().strftime("%Y%m%d")


if page == "Raw Data SPEC List":

    st.title("📘 Raw Data SPEC List")

    schema_file = st.file_uploader("Upload eCRF Schema", type=["xlsx", "xls"])

    if schema_file:
        sponsor, protocol_no = extract_protocol_no_from_filename(schema_file.name)
        db = build_rawdata_spec(schema_file)
        st.session_state["raw_spec"] = db
        st.session_state["folder_map_df"] = db["visit"]

        try:

            st.subheader("Variable List")
            if db["variable"].empty:
                st.error("⚠️ Variable List is empty")
            else:
                st.dataframe(db["variable"], use_container_width=True)

            st.subheader("Value List")
            if db["code"].empty:
                st.info("No Value List found")
            else:
                st.dataframe(db["code"], use_container_width=True)


            # =========================
            # Visit List
            # =========================
            st.subheader("Visit List")

            if db["visit"].empty:
                st.info("No Visit List found")

            else:

                tab_all, tab_crf = st.tabs(["Overall", "By CRF Domain"])

                # -------------------------
                # Tab 1: All (Folder)
                # -------------------------
                with tab_all:
                    st.dataframe(
                        db["visit"],
                        use_container_width=True
                    )

                # -------------------------
                # Tab 2: By CRF (SoA)
                # -------------------------
                with tab_crf:

                    if "domain_visit" not in db or db["domain_visit"].empty:
                        st.info("No Domain Visit mapping found from SoA")

                    else:
                        # Optional：讓使用者選 CRF
                        crf_list = sorted(db["domain_visit"]["STAT_DATASET"].dropna().unique())

                        selected_crf = st.selectbox(
                            "Select CRF Domain",
                            options=crf_list
                        )

                        df = db["domain_visit"][
                            db["domain_visit"]["STAT_DATASET"] == selected_crf
                        ].copy()

                        st.dataframe(df, use_container_width=True)


            # =========================
            # Download Raw Data SPEC
            # =========================
            try:

                sheet_dict = {
                    "Variable_List": db["variable"],
                    "Value_List": db["code"],
                    "Visit_List": db["visit"],
                    "Domain_Visit_List": db["domain_visit"] if "domain_visit" in db else None
                }

                excel_bytes = Export_excel(sheet_dict)

                # 檔名
                file_name = f"{sponsor}_{protocol_no}_Raw Data SPEC_{today_str}.xlsx"

                
                st.download_button(
                    label="📥 Download Raw Data SPEC (Excel)",
                    data=excel_bytes,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error("Failed to generate download file")
                st.text(str(e))



        except Exception:
            st.error("Schema parse error")
            st.text(traceback.format_exc())


# =========================
# Data Process
# =========================
if page == "Data Process":

    st.title("🔄 Data Process")


    # Input
    schema_file = st.file_uploader("Upload eCRF Schema", type=["xlsx", "xls"])
    zip_file = st.file_uploader("Upload Dataset ZIP", type=["zip"])

    system_vars_input = st.text_input(
        "System Variables (comma separated)",
        value="SUBJID,VISIT,VISITNUM,USUBJID,SITEID,STUDYID"
    )

    
    if schema_file:
        sponsor, protocol_no = extract_protocol_no_from_filename(schema_file.name)
        db = build_rawdata_spec(schema_file)
        st.session_state["raw_spec"] = db
        st.session_state["folder_map_df"] = db["visit"]


    if st.button("Run"):
        if not schema_file or not zip_file:
            st.error("Please upload both CRF schema and ZIP")
            st.stop()

        system_vars = [
            x.strip().upper()
            for x in system_vars_input.split(",")
            if x.strip()
        ]


        try:
            zip_bytes = zip_file.getvalue()
            result = process_raw_data(schema_file, zip_file, system_vars)

            # ✅ unpack
            schema_df = result["schema_df"]
            raw_df = result["raw_df"]
            compare_df = result["compare_df"]
            preview_map = result["raw_preview_map"]
            raw_dataset_map = result["raw_dataset_map"]
            merged_dataset_map = result["merged_dataset_map"]
            merge_log_df = result["merge_log_df"]
            visit_map_df = result["visit_map_df"]
            raw_spec = result["raw_spec"]

            # ✅ session
            st.session_state["schema"] = schema_df
            st.session_state["raw"] = raw_df
            st.session_state["compare"] = compare_df
            st.session_state["preview"] = preview_map
            st.session_state["raw_dataset_map"] = raw_dataset_map
            st.session_state["merged_dataset_map"] = merged_dataset_map
            st.session_state["merge_log_df"] = merge_log_df
            st.session_state["visit_map_df"] = visit_map_df

            st.success("Done ✅")

        except Exception as e:
            st.error("Error")
            st.text(traceback.format_exc())
            st.stop()

    st.divider()

    # =========================================================
    # UI
    # =========================================================
    if "compare" in st.session_state:

        compare_df = st.session_state["compare"]
        preview_map = st.session_state["preview"]
        raw_dataset_map = st.session_state["raw_dataset_map"]
        merged_dataset_map = st.session_state["merged_dataset_map"]
        merge_log_df = st.session_state["merge_log_df"]
        visit_map_df = st.session_state["visit_map_df"]



        st.subheader("📊 Issue Summary")
        render_summary(compare_df)


        # ✅ 主 tab
        tab1, tab2 = st.tabs(["📋 Compare List", "📦 Raw Data Review"])

        # =====================================================
        # Tab 1：總表
        # =====================================================
        with tab1:

            st.markdown("### 📋 Compare List")

            f1, f2, f3, f4 = st.columns([2, 2, 2, 1])

            dataset_filter = f1.multiselect(
                "Dataset",
                sorted(compare_df["RAW_DATASET"].dropna().unique())
            )

            status_filter = f2.multiselect(
                "Status",
                sorted(compare_df["STATUS"].dropna().unique())
            )

            keyword = f3.text_input("Keyword")

            only_issues = f4.checkbox("Only Issues", value=True)

            filtered_df = apply_filters(compare_df, dataset_filter, status_filter, keyword, only_issues)

            st.caption(f"{len(filtered_df)} / {len(compare_df)} rows")

            st.dataframe(style_compare(filtered_df), use_container_width=True)

            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "Download Filtered Compare CSV",
                    filtered_df.to_csv(index=False).encode("utf-8-sig"),
                    f"{sponsor}_{protocol_no}_Filtered Compare List_{today_str}.csv",
                    mime="text/csv"
                )
            with c2:
                st.download_button(
                    "Download Full Compare CSV",
                    compare_df.to_csv(index=False).encode("utf-8-sig"),
                    f"{sponsor}_{protocol_no}_Full Compare List_{today_str}.csv",
                    mime="text/csv"
                )



        # =====================================================
        # Tab 2：Raw Data Review（dataset-level）
        # =====================================================
        with tab2:

            st.markdown("### 📦 Raw Data Review")

            dataset_list = sorted(compare_df["RAW_DATASET"].dropna().unique())

            selected_dataset = st.selectbox("Select Dataset", dataset_list)

            preview = preview_map.get(selected_dataset)

            if preview is None or preview.empty:
                st.warning("No data")
            else:

                preview = reorder_preview_columns(preview)

                st.dataframe(preview, use_container_width=True)

                st.caption(f"Showing first {len(preview)} rows")

                st.markdown("##### Variable Overview")
                st.write(f"Total Variables: {len(preview.columns)}")
                st.code(", ".join(preview.columns))
            
                dataset_issue_df = compare_df[
                    compare_df["RAW_DATASET"] == selected_dataset
                ].copy()

                dataset_issue_df = dataset_issue_df[
                    ~dataset_issue_df["STATUS"].isin(["MATCH", "SYSTEM_VAR"])
                ]

                # 該 dataset 的 compare issues
                st.markdown("##### Dataset Compare Issues")

                if dataset_issue_df.empty:
                    st.success("✅ No issues found in this dataset")
                else:
                    st.dataframe(style_compare(dataset_issue_df), use_container_width=True)

        st.divider()

    
        # =====================================================
        # Raw Data Summary
        # =====================================================
        st.subheader("📊 Raw Data Summary")

        raw_dataset_map = st.session_state.get("raw_dataset_map")

        if raw_dataset_map is None or len(raw_dataset_map) == 0:
            st.info("No raw dataset available")
        else:

            summary_rows = []

            for name, df in raw_dataset_map.items():

                if df is None:
                    nobs = 0
                    nvars = 0
                    cols = []
                else:
                    nobs = len(df)
                    nvars = len(df.columns)
                    cols = df.columns

            # -----------------------------
            # 找 >200 字串長度欄位
            # -----------------------------
                long_cols = []

                if df is not None and not df.empty:
                    for col in cols:

                        s = df[col]

                        # 只檢查字串欄位
                        if not pd.api.types.is_numeric_dtype(s):

                            non_null = s.dropna()

                            if len(non_null) == 0:
                                continue

                            max_len = non_null.astype(str).map(len).max()

                            if pd.notna(max_len) and int(max_len) > 200:
                                long_cols.append(col)

                summary_rows.append({
                    "Dataset": name,
                    "Number of Variables": nvars,
                    "Number of Observations": nobs,
                    "Variable Length > 200": ", ".join(long_cols)
                })

            summary_df = pd.DataFrame(summary_rows)

            # obs=0 標灰底
            def highlight_empty(row):
                if row["Number of Observations"] == 0:
                    return ["background-color: lightgrey"] * len(row)
                else:
                    return [""] * len(row)

            st.dataframe(
                summary_df.style.apply(highlight_empty, axis=1),
                use_container_width=True
            )

        st.divider()


    
        # =====================================================
        # Visit Mapping List (Before Merge)
        # =====================================================
        st.subheader("📄 Visit Mapping List")

        visit_map_df["VISITNUM"] = pd.to_numeric(visit_map_df["VISITNUM"], errors="coerce")

        visit_map_df = visit_map_df.sort_values(
            by="VISITNUM",
            na_position="last"
        ).reset_index(drop=True)


        if visit_map_df.empty:
            st.info("No visit mapping result")
        else:

            # 小 summary（可選但我建議留）
            c1, c2, c3 = st.columns(3)

            total = len(visit_map_df)
            matched = (visit_map_df["SOURCE"] == "MATCHED").sum()
            schema_only = (visit_map_df["SOURCE"] == "SCHEMA_ONLY").sum()
            raw_only = (visit_map_df["SOURCE"] == "RAWDATA_ONLY").sum()

            c1.metric("Total", total)
            c2.metric("Matched", matched)
            c3.metric("Unmatched (Raw Data Only)", raw_only)

            # 主表
            
            val_display_cols = ["SOURCE", "VISIT", "VISIT_DVP", "FULL_TERM", "VISITNUM", "REPEAT_FOLDER"]
            
            visit_map_export = visit_map_df.copy()
            visit_map_export = visit_map_export[
                [c for c in val_display_cols if c in visit_map_export.columns]
            ]

            st.dataframe(visit_map_export, use_container_width=True)


        st.divider()


        # =====================================================
        # Merged Data
        # =====================================================
        st.subheader("🧩 Merged Dataset")

        if merged_dataset_map:

            merged_keys = sorted(list(merged_dataset_map.keys()))
            selected_merged = st.selectbox("Select Merged Dataset", merged_keys)

            merged_df = merged_dataset_map[selected_merged]

            # =========================
            # Merge Info
            # =========================
            merge_info = merge_log_df[
                merge_log_df["RESULT_DATASET"] == selected_merged
            ]

            subitems = []
            merge_mode = None
            key_vars = None

            if not merge_info.empty:
                row = merge_info.iloc[0]

                merge_mode = row["MERGE_MODE"]
                key_vars = row["KEY_VARS"]
                subitems = [x.strip() for x in str(row["SUBITEMS"]).split(",") if x.strip()]


                c1, c2, c3 = st.columns(3)

                c1.write("**Merged Mode**")
                c1.success(str(row["MERGE_MODE"]).capitalize())

                subitems = [x.strip() for x in str(row["SUBITEMS"]).split(",") if x.strip()]

                subitem_display = []

                for ds in subitems:
                    if ds in raw_dataset_map:
                        n = len(raw_dataset_map[ds])
                        subitem_display.append(f"{ds} (n={n})")
                    else:
                        subitem_display.append(ds)

                c2.write("**Sub-items Merged**")
                c2.info(", ".join(subitem_display))


                c3.write("**Key Variables**")
                c3.code(str(key_vars) if key_vars else "-")


            # =========================
            # Data Preview
            # =========================
            if merged_df is None or merged_df.empty:
                st.warning("Merged dataset is empty")

            else:
                view_df = prepare_view_df(merged_df, merge_mode)

                st.markdown("##### 📦 Review")
                st.dataframe(view_df.head(50), use_container_width=True)
                st.caption(f"Showing first {min(len(view_df), 50)} rows")

                st.write(f"Total Variables: {len(view_df.columns)}")
                st.code(", ".join(view_df.columns))

                st.write(f"Total Observations: {len(view_df)}")

                # 只有 vertical 才顯示來源分布
                if merge_mode == "vertical" and "SOURCEDN" in view_df.columns:
                    st.markdown("### Source Distribution")
                    source_dist = (
                        merged_df["SOURCEDN"]
                        .value_counts()
                        .rename_axis("SOURCEDN")
                        .reset_index(name="COUNT")
                    )
                    st.dataframe(source_dist, use_container_width=True)

        st.divider()

        # =====================================================
        # Export
        # =====================================================
        st.subheader("📤 Export")

        if "merged_dataset_map" in st.session_state:

            merged_dataset_map = st.session_state.get("merged_dataset_map")
            raw_df = st.session_state.get("raw")
            merge_log_df = st.session_state.get("merge_log_df")


            if merged_dataset_map is None or len(merged_dataset_map) == 0:
                st.info("No merged dataset available")

            else:
                export_dataset_map = {}

                for name, df in merged_dataset_map.items():
                    export_dataset_map[name] = prepare_view_df(df, merge_mode)


                # 顯示可輸出的 dataset
                export_names = [
                    name for name, df in export_dataset_map.items()
                    if df is not None
                ]

                export_format = st.radio(
                    "Select Export Format",
                    ["Excel (.xlsx)", "CSV ZIP (.zip)"],
                    horizontal=True,
                    key="export_format_radio"
                )


                st.write(f"Datasets to export: {len(export_names)}")
                if export_names:
                    st.code(", ".join(export_names))

                if export_format == "Excel (.xlsx)":
                    excel_bytes = build_merged_excel_bytes(export_dataset_map)

                    st.download_button(
                        label="📊 Download Datasets (Excel)",
                        data=excel_bytes,
                        file_name=f"{sponsor}_{protocol_no}_Rawdata_{today_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_merged_excel"
                    )

                    st.caption("All merged datasets are exported into one Excel workbook, one sheet per dataset.")

                else:          
                        
                    zip_bytes = build_csv_zip_bytes(
                        merged_dataset_map=export_dataset_map,
                        raw_df=raw_df,
                        merge_log_df=merge_log_df
                    )
     
                    st.download_button(
                        label="📦 Download CSV + SAS Script (ZIP)",
                        data=zip_bytes,
                        file_name=f"{sponsor}_{protocol_no}_Rawdata_{today_str}.zip",
                        mime="application/zip",
                        use_container_width=True,
                        key="download_csv_zip"
                    )

                    st.caption("All merged datasets are exported in CSV files inside one ZIP archive.")






# =========================
# Data Check
# =========================
if page == "DVP Review and Data Check":

    st.title("🔍 DVP Review and Data Check")

    st.header("Step 1 | DVP Review")

    # Input
    schema_file = st.file_uploader("Upload eCRF Schema", type=["xlsx", "xls"])
    dvp_file = st.file_uploader("Upload DVP", type=["xlsx", "xls"])

 
    # =========================
    # Run Step 1
    # =========================
    if st.button("Run Step 1", key="run_dvp_review"):

        if schema_file:
            sponsor, protocol_no = extract_protocol_no_from_filename(schema_file.name)
            st.session_state["sponsor"] = sponsor
            st.session_state["protocol_no"] = protocol_no


        if not schema_file or not dvp_file:
            st.error("Please upload both eCRF Schema and DVP file.")
            st.stop()

        try:
            # 1) Parse schema
            raw_spec = build_rawdata_spec(schema_file)
            st.session_state["raw_spec"] = raw_spec

            

            domain_visit_df = raw_spec["domain_visit"]

            # 2) Parse DVP
            dvp_df = parse_dvp_file(dvp_file, visit_df=domain_visit_df)
            st.session_state["dvp_df"] = dvp_df

            # 3) Review DVP
            dvp_review_df = build_dvp_review_df(
                dvp_df=dvp_df,
                raw_spec=raw_spec
            )

            st.session_state["dvp_review_df"] = dvp_review_df

            # Clear previous Step 2 results when Step 1 reruns
            for k in [
                "dvp_result_df",
                "dvp_fail_detail_map",
                "dvp_enriched_dataset_map",
                "dvp_merge_debug_df",
                "dvp_merged_dataset_map",
            ]:
                if k in st.session_state:
                    del st.session_state[k]


        except KeyError as e:
            st.error(f"❌ Missing required schema output or column: {e}")

        except Exception as e:
            st.error(f"❌ Error while processing DVP review: {e}")


    st.divider()

    # =========================================================
    # Step 1 UI
    # =========================================================
    if "dvp_review_df" in st.session_state:

        dvp_review_df = st.session_state["dvp_review_df"].copy()
        dvp_df = st.session_state.get("dvp_df", pd.DataFrame())

        if dvp_review_df.empty:
            st.warning("No DVP review rows generated.")
        else:

            # =========================
            # Top Summary
            # =========================
            st.markdown("### 📊 Summary")

            df = dvp_review_df.copy()

            # 保險欄位
            if "UNKNOWN_VARS" not in df.columns:
                df["UNKNOWN_VARS"] = [[] for _ in range(len(df))]

            if "INVALID_VISIT_LABELS" not in df.columns:
                df["INVALID_VISIT_LABELS"] = [[] for _ in range(len(df))]

            if "NEED_UNIT_STANDARDIZATION" not in df.columns:
                df["NEED_UNIT_STANDARDIZATION"] = False

            if "NEED_REFERENCE_RANGE" not in df.columns:
                df["NEED_REFERENCE_RANGE"] = False

            if "READY_FOR_STEP2" not in df.columns:
                df["READY_FOR_STEP2"] = False

            if "EXECUTION_CLASS" not in df.columns:
                df["EXECUTION_CLASS"] = ""

            if "FORM" not in df.columns:
                df["FORM"] = ""

            if "STAT_DATASET" not in df.columns:
                df["STAT_DATASET"] = ""

            if "RULE" not in df.columns:
                df["RULE"] = ""

            if "EFFECTIVE_VISIT" not in df.columns:
                df["EFFECTIVE_VISIT"] = pd.NA

            # helper flags
            df["_HAS_UNKNOWN"] = df["UNKNOWN_VARS"].apply(
                lambda x: len(x) > 0 if isinstance(x, list)
                else bool(str(x).strip() not in ["", "[]", "nan", "None"])
            )

            df["_HAS_INVALID_VISIT"] = df["INVALID_VISIT_LABELS"].apply(
                lambda x: len(x) > 0 if isinstance(x, list)
                else bool(str(x).strip() not in ["", "[]", "nan", "None"])
            )

            df["_IS_EXECUTABLE"] = df["READY_FOR_STEP2"].astype(bool)
            df["_IS_SKIP"] = ~df["_IS_EXECUTABLE"]

            # =========================
            # Original DVP Summary
            # =========================
            raw_rule_n = (
                dvp_df["RULE"]
                .dropna()
                .astype(str)
                .str.strip()
                .nunique()
                if not dvp_df.empty and "RULE" in dvp_df.columns
                else 0
            )       

            # =========================
            # Execution Outcome
            # 分類：Total = Executable + Skipped
            # =========================
            total_blocks = len(df)
            executable_n = int(df["_IS_EXECUTABLE"].sum())
            skip_n = int(df["_IS_SKIP"].sum())

            e1, e2, e3, e4 = st.columns(4)
            e1.metric("Unique Rules", raw_rule_n)
            e2.metric("Total Expanded Rules", total_blocks)
            e3.metric("Executable", executable_n)
            e4.metric("Skipped", skip_n)

            # =========================
            # Skip Reason Breakdown
            # 可重疊原因，只針對 skipped blocks 統計
            # =========================
            st.markdown("###### Skip Reason Breakdown")
            skip_df = df[df["_IS_SKIP"]].copy()

            need_unit_n = int(
                skip_df["NEED_UNIT_STANDARDIZATION"].astype(bool).sum()
            )

            need_range_n = int(
                skip_df["NEED_REFERENCE_RANGE"].astype(bool).sum()
            )

            unknown_n = int(
                skip_df["_HAS_UNKNOWN"].sum()
            )

            invalid_visit_n = int(
                skip_df["_HAS_INVALID_VISIT"].sum()
            )

            r1, r2, r3, r4 = st.columns(4)

            r1.metric("Need Unit", need_unit_n)
            r2.metric("Need Range", need_range_n)
            r3.metric("Unknown Var", unknown_n)
            r4.metric("Invalid Visit", invalid_visit_n)
            st.caption("Skip reasons may overlap. One expended-rule can have multiple reasons.")

            st.markdown("""
            **Execution Class 說明**
            - ✅ `EXECUTABLE`：可進 Step 2 執行 Data Check
            - ⚠️ `NEED_UNIT_STANDARDIZATION`：條件含單位或比例
            - ⚠️ `NEED_REFERENCE_RANGE`：含 ULN / LLN
            - ❌ `UNRESOLVED_SCHEMA`：schema 找不到變數
            - ❌ `INVALID_VISIT_LABEL`：visit label 不存在於 schema
            """)

            # =========================
            # Summary by Domain
            # =========================
            st.markdown("### - by Domain")

            summary_df = (
                df
                .groupby(["STAT_DATASET", "FORM"], dropna=False)
                .agg(
                    RULE_N=("RULE", "nunique"),

                    EXPANDED_RULE_N=("RULE", "count"),

                    EXECUTABLE_N=("_IS_EXECUTABLE", "sum"),

                    INVALID_VISIT_LABEL=("_HAS_INVALID_VISIT", "sum"),

                    NEED_UNIT=("NEED_UNIT_STANDARDIZATION", lambda s: int(s.astype(bool).sum())),

                    NEED_RANGE=("NEED_REFERENCE_RANGE", lambda s: int(s.astype(bool).sum())),

                    UNKNOWN=("_HAS_UNKNOWN", "sum"),

                    SKIP=("_IS_SKIP", "sum"),
                )
                .reset_index()
            )

            summary_cols = [
                "STAT_DATASET",
                "FORM",
                "RULE_N",
                "EXPANDED_RULE_N",
                "EXECUTABLE_N",
                "SKIP",
                "INVALID_VISIT_LABEL",
                "NEED_UNIT",
                "NEED_RANGE",
                "UNKNOWN",

            ]

            summary_cols = [c for c in summary_cols if c in summary_df.columns]

            st.dataframe(
                summary_df[summary_cols],
                use_container_width=True
            )

            st.divider()

            # =========================
            # DVP Detail Table
            # =========================
            st.markdown("### DVP Review Detail")

            show_df = df.copy()

            show_df["HAS_UNIT"] = show_df["NEED_UNIT_STANDARDIZATION"]
            show_df["HAS_RANGE"] = show_df["NEED_REFERENCE_RANGE"]

            # Filters
            f1, f2, f3, f4 = st.columns([2, 2, 2, 3])

            stat_filter = f1.multiselect(
                "STAT Dataset",
                sorted(show_df["STAT_DATASET"].dropna().astype(str).unique()), key="step1_stat_filter"
            )

            type_filter = f2.multiselect(
                "Rule Type",
                sorted(show_df["RULE_TYPE"].dropna().astype(str).unique())
                if "RULE_TYPE" in show_df.columns else [], key="step1_type_filter"
            )

            exec_filter = f3.multiselect(
                "Execution Class",
                sorted(show_df["EXECUTION_CLASS"].dropna().astype(str).unique()), key="step1_exec_filter"
            )

            keyword = f4.text_input(
                "Keyword",
                key="dvp_review_keyword"
            )

            if stat_filter:
                show_df = show_df[
                    show_df["STAT_DATASET"].astype(str).isin(stat_filter)
                ]

            if type_filter and "RULE_TYPE" in show_df.columns:
                show_df = show_df[
                    show_df["RULE_TYPE"].astype(str).isin(type_filter)
                ]

            if exec_filter:
                show_df = show_df[
                    show_df["EXECUTION_CLASS"].astype(str).isin(exec_filter)
                ]

            if keyword:
                kw = keyword.upper().strip()
                show_df = show_df[
                    show_df.apply(
                        lambda r: kw in str(r.to_dict()).upper(),
                        axis=1
                    )
                ]

            review_cols = [
                "RULE",
                "STAT_DATASET",
                "FORM",
                "BLOCK_ID",
                "EFFECTIVE_VISIT",

                "RULE_TYPE",
                "EXECUTION_CLASS",
                "READY_FOR_STEP2",
                "BLOCKING_REASON",

                "HAS_UNIT",
                "HAS_RANGE",

                "VISIT_RAW",
                "DVP_VISIT_SCOPE",
                "BLOCK_VISIT_SCOPE",
                "INVALID_VISIT_LABELS",
                "EFFECTIVE_VISIT_SCOPE",

                "TARGET_VARS",
                "REF_DATASETS",
                "REF_VARS",
                "UNKNOWN_VARS",
                "RULE_VAR_TOKENS",
                "AMBIGUOUS_REFS",

                "PARAMS",
                "PARAM_STRATEGY",

                "MESSAGE",
                "CONDITION_BLOCK",
            ]

            review_cols = [c for c in review_cols if c in show_df.columns]

            st.caption(f"{len(show_df)} / {len(df)} expanded blocks")
            st.dataframe(show_df[review_cols], use_container_width=True)

            # =========================
            # Issues / Not Executable
            # =========================
            st.markdown("### Issues / Not Executable")

            issue_df = df[
                (df["_HAS_UNKNOWN"])
                | (df["_HAS_INVALID_VISIT"])
                | (~df["READY_FOR_STEP2"].astype(bool))
            ].copy()

            if issue_df.empty:
                st.success("No issues ✅")

            else:
                issue_df["HAS_UNIT"] = issue_df["NEED_UNIT_STANDARDIZATION"]
                issue_df["HAS_RANGE"] = issue_df["NEED_REFERENCE_RANGE"]

                issue_cols = [c for c in review_cols if c in issue_df.columns]

                st.warning(f"{len(issue_df)} problematic expanded rules")
                st.dataframe(issue_df[issue_cols], use_container_width=True)


            # =========================
            # Download DVP Review Excel
            # =========================
            detail_df = dvp_review_df.copy()

            sheet_dict = {
                "DVP Review": dvp_review_df,
            }

            excel_bytes = Export_excel(sheet_dict)
            
            sponsor = st.session_state.get("sponsor", "UNKNOWN_SPONSOR")
            protocol_no = st.session_state.get("protocol_no", "UNKNOWN_PROTOCOL")

            file_name = f"{sponsor}_{protocol_no}_DVP_Review Detail_{today_str}.xlsx"

            st.download_button(
                label="📥 Download DVP Review Detail (Excel)",
                data=excel_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


            st.divider()

    # =========================================================
    # Step 2 | Data Check
    # =========================================================

    if "dvp_review_df" in st.session_state:

        st.header("Step 2 | Data Check")

        zip_file = st.file_uploader(
            "Upload SAS ZIP",
            type=["zip"],
            key="dvp_zip_file"
        )

        system_vars_input = st.text_input(
            "System Variables",
            value="SUBJID,VISIT,VISITNUM,USUBJID,SITEID,STUDYID",
            key="dvp_system_vars"
        )

        if st.button("Run Step 2", key="run_dvp_data_check"):

            if not zip_file:
                st.error("Please upload SAS ZIP for Step 2.")
                st.stop()

            if "raw_spec" not in st.session_state:
                st.error("raw_spec is missing. Please rerun Step 1.")
                st.stop()

            try:
                raw_spec = st.session_state["raw_spec"]
                dvp_review_df = st.session_state["dvp_review_df"]

                system_vars = [
                    x.strip().upper()
                    for x in system_vars_input.split(",")
                    if x.strip()
                ]

                # 1) Process raw SAS zip
                process_result = process_raw_data(
                    schema_file,
                    zip_file,
                    system_vars
                )

                merged_dataset_map = process_result["merged_dataset_map"]

                # 2) Run Data Check v2
                result_df, fail_detail_map, enriched_dataset_map, merge_debug_df = run_dvp_data_check(
                    dvp_review_df=dvp_review_df,
                    merged_dataset_map=merged_dataset_map,
                    raw_spec=raw_spec
                )

                st.session_state["dvp_result_df"] = result_df
                st.session_state["dvp_fail_detail_map"] = fail_detail_map
                st.session_state["dvp_enriched_dataset_map"] = enriched_dataset_map
                st.session_state["dvp_merge_debug_df"] = merge_debug_df
                st.session_state["dvp_merged_dataset_map"] = merged_dataset_map

                st.success("Step 2 completed ✅")

            except Exception as e:
                st.error(f"❌ Error while running Data Check: {e}")
                st.text(traceback.format_exc())


    # =========================================================
    # Step 2 Result UI
    # =========================================================
    if "dvp_result_df" in st.session_state:

        result_df = st.session_state["dvp_result_df"].copy()
        fail_detail_map = st.session_state.get("dvp_fail_detail_map", {})
        enriched_dataset_map = st.session_state.get("dvp_enriched_dataset_map", {})
        merge_debug_df = st.session_state.get("dvp_merge_debug_df", pd.DataFrame())

        st.divider()
        st.markdown("### 📊 Execution Summary")

        if result_df.empty:
            st.warning("No Data Check results.")
        else:

            # -------------------------------------------------
            # Metrics
            # -------------------------------------------------
            c1, c2, c3, c4, c5 = st.columns(5)

            c1.metric("Total Blocks", len(result_df))

            c2.metric(
                "OK",
                int((result_df["STATUS"] == "OK").sum())
                if "STATUS" in result_df.columns else 0
            )

            c3.metric(
                "Eval Error",
                int(
                    result_df["STATUS"]
                    .astype(str)
                    .str.startswith("EVAL_ERROR")
                    .sum()
                )
                if "STATUS" in result_df.columns else 0
            )

            c4.metric(
                "Skipped",
                int((result_df["STATUS"] == "SKIP_STEP2").sum())
                if "STATUS" in result_df.columns else 0
            )

            c5.metric(
                "Fail Rows",
                int(
                    pd.to_numeric(
                        result_df.get("FAIL_COUNT", 0),
                        errors="coerce"
                    )
                    .fillna(0)
                    .sum()
                )
            )

            # -------------------------------------------------
            # Execution Summary
            # -------------------------------------------------
            if "EXECUTION_CLASS" in result_df.columns:
                result_exec_summary = (
                    result_df
                    .groupby(["STATUS", "EXECUTION_CLASS"])
                    .size()
                    .reset_index(name="COUNT")
                )

                st.dataframe(result_exec_summary, use_container_width=True)

            # -------------------------------------------------
            # Ref Merge Summary
            # -------------------------------------------------
            st.markdown("### Ref Merge Summary")

            if merge_debug_df is None or merge_debug_df.empty:
                st.info("No ref merge debug available.")
            else:
                st.dataframe(merge_debug_df, use_container_width=True)

            # -------------------------------------------------
            # Result filters
            # -------------------------------------------------
            st.markdown("### Data Check Status")

            rf1, rf2, rf3, rf4 = st.columns([2, 2, 2, 3])

            status_filter = rf1.multiselect(
                "Status",
                sorted(result_df["STATUS"].dropna().astype(str).unique())
                if "STATUS" in result_df.columns else [], key="step2_status_filter"
            )

            stat_filter2 = rf2.multiselect(
                "STAT Dataset",
                sorted(result_df["STAT_DATASET"].dropna().astype(str).unique())
                if "STAT_DATASET" in result_df.columns else [], key="step2_stat_filter"
            )

            exec_filter2 = rf3.multiselect(
                "Execution Class",
                sorted(result_df["EXECUTION_CLASS"].dropna().astype(str).unique())
                if "EXECUTION_CLASS" in result_df.columns else [], key="step2_exec_filter"
            )

            keyword2 = rf4.text_input(
                "Keyword",
                key="dvp_result_keyword"
            )

            show_result = result_df.copy()

            if status_filter and "STATUS" in show_result.columns:
                show_result = show_result[
                    show_result["STATUS"].astype(str).isin(status_filter)
                ]

            if stat_filter2 and "STAT_DATASET" in show_result.columns:
                show_result = show_result[
                    show_result["STAT_DATASET"].astype(str).isin(stat_filter2)
                ]

            if exec_filter2 and "EXECUTION_CLASS" in show_result.columns:
                show_result = show_result[
                    show_result["EXECUTION_CLASS"].astype(str).isin(exec_filter2)
                ]

            if keyword2:
                kw = keyword2.upper().strip()
                show_result = show_result[
                    show_result.apply(
                        lambda r: kw in str(r.to_dict()).upper(),
                        axis=1
                    )
                ]

            result_cols = [
                "RULE",
                "STAT_DATASET",
                "BLOCK_ID",
                "RULE_TYPE",
                "EXECUTION_CLASS",
                "STATUS",
                "FAIL_COUNT",
                "RAW_FAIL_COUNT",
                "IS_ITEM_LEVEL",
                "FAIL_GRAIN_KEYS",

                "VISIT_SCOPE",
                "TARGET_VARS",
                "REF_DATASETS",
                "REF_VARS",

                "BLOCKING_REASON",
                "MESSAGE",
                "CONDITION_BLOCK",
                "PYTHON_EXPR",
            ]

            result_cols = [c for c in result_cols if c in show_result.columns]

            st.caption(f"{len(show_result)} / {len(result_df)} result rows")
            st.dataframe(show_result[result_cols], use_container_width=True)



            # -------------------------------------------------
            # Skipped / Not Executable Rules
            # -------------------------------------------------
            st.markdown("### Skipped / Not Executable Rules")

            skip_df = result_df[
                result_df["STATUS"].astype(str).eq("SKIP_STEP2")
            ].copy()

            if skip_df.empty:
                st.success("No skipped rules.")
            else:
                skip_cols = [
                    "RULE",
                    "STAT_DATASET",
                    "BLOCK_ID",
                    "RULE_TYPE",
                    "EXECUTION_CLASS",
                    "BLOCKING_REASON",
                    "CONDITION_BLOCK",
                    "TARGET_VARS",
                    "REF_VARS",
                ]

                skip_cols = [c for c in skip_cols if c in skip_df.columns]

                st.dataframe(skip_df[skip_cols], use_container_width=True)

            st.divider()
            
            # -------------------------------------------------
            # Fail Detail
            # -------------------------------------------------
            st.markdown("### Fail Detail")

            failed_df = result_df[
                pd.to_numeric(
                    result_df.get("FAIL_COUNT", 0),
                    errors="coerce"
                )
                .fillna(0)
                > 0
            ].copy()

            if failed_df.empty:
                st.success("No fail rows ✅")
            else:
                failed_df["LABEL"] = (
                    failed_df["RULE"].astype(str)
                    + " | "
                    + failed_df["STAT_DATASET"].astype(str)
                    + " | Block "
                    + failed_df["BLOCK_ID"].astype(str)
                    + " | Visit "
                    + failed_df["EFFECTIVE_VISIT"].fillna("ALL").astype(str)
                    + " | n="
                    + failed_df["FAIL_COUNT"].astype(str)
                    + " | ID="
                    + failed_df["IDX"].astype(str)
                )

                failed_df = result_df[
                    pd.to_numeric(
                        result_df.get("FAIL_COUNT", 0),
                        errors="coerce"
                    )
                    .fillna(0)
                    > 0
                ].copy()

                if failed_df.empty:
                    st.success("No fail rows ✅")
                else:
                    # build dropdown 前：排序資料
                    failed_df = (
                        failed_df
                        .assign(
                            STAT_SORT=lambda d: d["STAT_DATASET"].astype(str),
                            RULE_SORT=lambda d: d["RULE"].astype(str),
                            VISIT_SORT=lambda d: d["EFFECTIVE_VISIT"].astype(str),
                        )
                        .sort_values(
                            by=[
                                "STAT_SORT",
                                "RULE_SORT",
                                "BLOCK_ID",
                                "VISIT_SORT",
                                "IDX",
                            ],
                            ascending=[True, True, True, True, True],
                            na_position="last"
                        )
                        .reset_index(drop=True)
                    )


                    # LABEL（排序後再產）
                    failed_df["LABEL"] = (
                        failed_df["STAT_DATASET"].astype(str)
                        + " | "
                        + failed_df["RULE"].astype(str)
                        + " | Block "
                        + failed_df["BLOCK_ID"].astype(str)
                        + " | Visit "
                        + failed_df["EFFECTIVE_VISIT"].fillna("ALL").astype(str)
                        + " | n="
                        + failed_df["FAIL_COUNT"].astype(str)
                    )

                    # =====================================================
                    # dropdown（用排序後的）
                    # =====================================================
                    selected_label = st.selectbox(
                        "Select failed rule/block",
                        failed_df["LABEL"].tolist(),
                        key="select_failed_rule"
                    )

                    # =====================================================
                    # 取回 row（要用排序後 df）
                    # =====================================================
                    selected_row = failed_df.loc[
                        failed_df["LABEL"] == selected_label
                    ].iloc[0]

                    selected_idx = selected_row["IDX"]
                    detail = fail_detail_map.get(selected_idx, {})


                st.write("**Message**")
                st.info(str(detail.get("MESSAGE", "")))

                st.write("**Condition**")
                st.code(str(detail.get("CONDITION_BLOCK", "")), language="text")

                st.write("**Python Expression**")
                st.code(str(detail.get("PYTHON_EXPR", "")), language="python")

                st.write("**Target / Ref Variables**")
                st.json({
                    "TARGET_VARS": detail.get("TARGET_VARS", []),
                    "REF_DATASETS": detail.get("REF_DATASETS", []),
                    "REF_VARS": detail.get("REF_VARS", []),
                    "VISIT_SCOPE": detail.get("VISIT_SCOPE", []),
                })

                fail_df = detail.get("FAIL_DF", pd.DataFrame())

                if fail_df is None or fail_df.empty:
                    st.success("No fail rows for this selected rule.")
                else:
                    st.dataframe(fail_df, use_container_width=True)


                with st.expander("Eval Dataset Preview", expanded=False):
                    eval_df = detail.get("EVAL_DF", pd.DataFrame())

                    if eval_df is None or eval_df.empty:
                        st.info("No eval dataset for this rule.")
                    else:
                        st.dataframe(eval_df.head(100), use_container_width=True)
                        st.caption(f"Rows: {len(eval_df)}, Columns: {len(eval_df.columns)}")

            # -------------------------------------------------
            # Enriched Dataset Preview
            # -------------------------------------------------
            st.markdown("### Enriched Dataset Preview")

            if not enriched_dataset_map:
                st.info("No enriched dataset available.")
            else:
                ds_names = sorted(enriched_dataset_map.keys())

                selected_ds = st.selectbox(
                    "Select enriched STAT_DATASET",
                    ds_names,
                    key="select_enriched_dataset"
                )

                preview_df = enriched_dataset_map[selected_ds]

                st.dataframe(
                    preview_df,
                    use_container_width=True
                )

                st.caption(
                    f"Rows: {len(preview_df)}, Columns: {len(preview_df.columns)}"
                )


            st.markdown("### Unified Failed Report")

            report_df = build_unified_fail_report_df(
                result_df=result_df,
                fail_detail_map=fail_detail_map,
                raw_spec=st.session_state.get("raw_spec", {})
            )

            if report_df.empty:
                st.success("No failed records ✅")
            else:

                st.dataframe(report_df, use_container_width=True)

                # =========================
                # Download DVP Review Excel
                # =========================
                sheet_dict = {
                    "Data Check Result": report_df,
                }

                excel_bytes = Export_excel(sheet_dict)
            
                sponsor = st.session_state.get("sponsor", "UNKNOWN_SPONSOR")
                protocol_no = st.session_state.get("protocol_no", "UNKNOWN_PROTOCOL")

                file_name = f"{sponsor}_{protocol_no}_Data Check Result_{today_str}.xlsx"

                st.download_button(
                    label="📥 Download Data Check Result (Excel)",
                    data=excel_bytes,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )



if page == "Verify":

    st.header("Verify")

    col1, col2 = st.columns(2)

    with col1:
        bs_task = st.selectbox(
            "BS Task",
            [
                "Listing",
                "Table",
                "Raw Data",
                "SDTM",
                "ADaM",
                "SAS Check",
            ],
        )

    with col2:
        compare_type = st.selectbox(
            "Compare Type",
            [
                "Main/Double Verification",
                "Version Comparison",
            ],
        )

    if compare_type == "Main/Double Verification":
        left_label = "Main"
        right_label = "Double"
        upload_left_label = "Upload Main File"
        upload_right_label = "Upload Double File"
        help_text = (
            f"Use this mode for Main vs Double programming verification of {bs_task}."
        )
    else:
        left_label = "New"
        right_label = "Previous"
        upload_left_label = "Upload New File"
        upload_right_label = "Upload Previous File"
        help_text = (
            f"Use this mode for current/new output vs previous version comparison of {bs_task}."
        )

    st.info(help_text)

    st.markdown(
        f"""
        **Verify Note**

        1. Check sheet consistency first. Sheet differences are summarized in `Summary`.
        2. Sheets existing only in `{left_label}` or `{right_label}` are skipped.
        3. For common sheets, columns are compared by variable name / label.
        4. Columns existing only in one side are skipped for data comparison.
        """
    )

    # Upload files
    col_left, col_right = st.columns(2)

    with col_left:
        file_left = st.file_uploader(
            upload_left_label,
            type=["xlsx", "xls"],
            key=f"verify_file_left_{compare_type}",
        )

    with col_right:
        file_right = st.file_uploader(
            upload_right_label,
            type=["xlsx", "xls"],
            key=f"verify_file_right_{compare_type}",
        )

    # --------------------------------------------------
    # Sponsor--# Sponsor / Protocol
    detected_sponsor = ""
    detected_protocol = ""

    if file_left is not None:

        detected_sponsor, detected_protocol = (
            extract_protocol_no_from_filename(
                file_left.name
            )
        )

    if (
        detected_sponsor
        and not st.session_state.get("sponsor")
    ):
        st.session_state["sponsor"] = detected_sponsor

    if (
        detected_protocol
        and not st.session_state.get("protocol_no")
    ):
        st.session_state["protocol_no"] = detected_protocol


    col_sp1, col_sp2 = st.columns(2)

    with col_sp1:
        sponsor_manual = st.text_input(
            "Sponsor",
            key="sponsor",
        )

    with col_sp2:
        protocol_manual = st.text_input(
            "Protocol No.",
            key="protocol_no",
            help=(
                "If protocol number cannot be identified "
                "from filename (e.g. SDTM / ADaM), "
                "please enter manually."
            ),
        )

    run_verify = st.button(
        "Run Verify",
        type="primary",
        use_container_width=True,
    )

    if run_verify:

        if file_left is None or file_right is None:
            st.error(f"Please upload both {left_label} and {right_label} Excel files.")

        else:

            st.session_state["verify_result"] = compare_excel_verify(
                file_left=file_left,
                file_right=file_right,
                compare_type=compare_type,
            )

            st.session_state["verify_result_compare_type"] = compare_type

    result = st.session_state.get("verify_result")

    if result:

        result_compare_type = st.session_state.get(
            "verify_result_compare_type",
            compare_type,
        )

        result_labels = result.get("Labels", {})
        result_left_label = result_labels.get("LEFT", left_label)
        result_right_label = result_labels.get("RIGHT", right_label)

        st.success(
            f"Verify result is available: {result_left_label} vs {result_right_label}"
        )


        # Summary
        summary_df = result.get("Summary", pd.DataFrame())
        sheet_reports = result.get("Sheet_Reports", {})

        summary_view = build_summary_view(result)
        all_sheets = summary_view["Sheet"].tolist()

        st.subheader("Summary")

        if summary_df.empty:
            st.info("No summary result.")
        else:
            summary_view = build_summary_view(result)
            all_sheets = summary_view["Sheet"].tolist()

            st.dataframe(
                style_summary(summary_view),
                use_container_width=True,
                hide_index=True,
            )


        st.subheader("Data Difference by Sheet")

        summary_view = build_summary_view(result)
        sheet_options = summary_view["Sheet"].tolist()

        selected_sheet = st.selectbox(
            "Sheet",
            sheet_options,
        )
        
        structure_info = get_structure_diff_summary(
            structure_df=result.get(
                "Structure_Diff",
                pd.DataFrame(),
            ),
            sheet_name=selected_sheet,
            labels=result["Labels"],
            summary_df=result.get("Summary", pd.DataFrame(),),
        )

        if (
            structure_info["left_only"]
            or structure_info["right_only"]
        ):

            left = result["Labels"]["LEFT"]
            right = result["Labels"]["RIGHT"]

            st.markdown("**Structure Difference**")

            st.write(
                f"{left} Only: "
                f"{structure_info['left_only'] or '-'}"
            )   

            st.write(
                f"{right} Only: "
                f"{structure_info['right_only'] or '-'}"
            )

        
        if structure_info["sheet_only"]:

            st.warning(
                f"Sheet Difference: "
                f"{structure_info['sheet_only']}"
            )

        df = sheet_reports.get(
            selected_sheet,
            pd.DataFrame(),
        )

        has_structure_diff = (
            structure_info["sheet_only"]
            or structure_info["right_only"] 
            or structure_info["left_only"]
        )

        if df.empty and not has_structure_diff:
            st.success(
                f"No data difference in sheet: {selected_sheet}"
            )
        else:
            if not df.empty:
                st.dataframe(
                    style_data_diff(df, result),
                    use_container_width=True,
                    hide_index=True,
                )

        # Download
        report_file = export_verify_report(result)

        sponsor = sponsor_manual.strip()

        protocol_no = (
            protocol_manual.strip()
            or "UNKNOWN_PROTOCOL"
        )

        if sponsor:
            file_name = (
                f"{sponsor}_{protocol_no}"
                f"_Verify {bs_task} Result_{today_str}.xlsx"
            )

        else:
            file_name = (
                f"{protocol_no}"
                f"_Verify {bs_task} Result_{today_str}.xlsx"
            )


        st.download_button(
            label="Download Verify Result",
            data=report_file,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
