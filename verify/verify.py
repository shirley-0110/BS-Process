import io
import xlrd
from collections import Counter, defaultdict
from datetime import date, datetime, time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter




# =========================================================
# 偵測Header
# =========================================================
def _is_empty_cell(value):
    if value is None:
        return True

    if isinstance(value, str) and value.strip() == "":
        return True

    return False


def _detect_header_row(ws, max_scan_rows=50):
    """
    Auto-detect header row by the first non-empty cell in column A.

    Rule:
    - If sheet name is TOC, use row 1 as header.
    - Otherwise, scan column A from row 1 downward.
    - The first row where column A is not empty is treated as header.
    """

    max_row = ws.max_row or 0

    if max_row == 0:
        return 1

    # TOC sheet always uses first row as header
    if str(ws.title).strip().upper() == "TOC":
        return 1

    scan_until = min(max_row, max_scan_rows)

    for row_idx in range(1, scan_until + 1):
        first_col_value = ws.cell(row=row_idx, column=1).value

        if not _is_empty_cell(first_col_value):
            return row_idx

    # fallback:
    # if column A is empty within scan range, continue scanning all rows
    for row_idx in range(scan_until + 1, max_row + 1):
        first_col_value = ws.cell(row=row_idx, column=1).value

        if not _is_empty_cell(first_col_value):
            return row_idx

    # if column A is entirely empty, fallback to row 1
    return 1





# =========================================================
# Compare labels
# =========================================================

def _get_compare_labels(compare_type):
    """
    Visible labels used in UI/report.

    Internal logic still compares left/right files,
    but all report-visible names use workflow labels.
    """
    if compare_type == "Double Programming":
        return {
            "LEFT": "Main",
            "RIGHT": "Double",
            "ONLY_LEFT": "Only in Main",
            "ONLY_RIGHT": "Only in Double",
            "BOTH": "Main / Double",
        }

    if compare_type == "Version Comparison":
        return {
            "LEFT": "New",
            "RIGHT": "Previous",
            "ONLY_LEFT": "Only in New",
            "ONLY_RIGHT": "Only in Previous",
            "BOTH": "New / Previous",
        }

    return {
        "LEFT": "Main",
        "RIGHT": "Double",
        "ONLY_LEFT": "Only in Main",
        "ONLY_RIGHT": "Only in Double",
        "BOTH": "Main / Double",
    }


# =========================================================
# Basic cell handling
# =========================================================

def _cell_token(value):
    """
    Exact token for comparison.
    No strip / no normalization / no case conversion.
    """
    if value is None:
        return "<None>"

    return f"<{type(value).__name__}>:{repr(value)}"


def _cell_display(value):
    """
    Display value in report.
    """
    if value is None:
        return ""

    if isinstance(value, (datetime, date, time)):
        return value.isoformat()

    return value


def _safe_sheet_name(name, existing_names):
    """
    Excel sheet name max length is 31.
    Also avoid duplicate sheet names.
    """
    base = str(name)[:31]
    candidate = base

    i = 1
    while candidate in existing_names:
        suffix = f"_{i}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        i += 1

    existing_names.add(candidate)
    return candidate


# =========================================================
# Excel read
# =========================================================

def _read_excel(file_obj, header_row=None):
    """
    Read all sheets by openpyxl. (Support: xlsx, xls)

    """

    file_name = getattr(
        file_obj,
        "name",
        "",
    ).lower()
    
    workbook_data = {}


    # ==================================================
    # XLS
    # ==================================================
    if file_name.endswith(".xls"):

        excel_file = pd.ExcelFile(
            file_obj,
            engine="xlrd",
        )

        for sheet_name in excel_file.sheet_names:

            df = pd.read_excel(
                excel_file,
                sheet_name=sheet_name,
                header=None,
                engine="xlrd",
            )

            max_row = len(df)
            max_col = len(df.columns)

            if max_row == 0 or max_col == 0:

                workbook_data[sheet_name] = {
                    "headers": [],
                    "rows": [],
                    "detected_header_row": "",
                    "max_row": max_row,
                    "max_col": max_col,
                }

                continue

            values = df.where(
                pd.notna(df),
                None,
            ).values.tolist()

            # ------------------------------------------
            # Detect header row
            # ------------------------------------------
            if header_row is not None:

                sheet_header_row = header_row

            else:

                sheet_header_row = 1

                if sheet_name.strip().upper() != "TOC":

                    for row_idx, row in enumerate(
                        values,
                        start=1,
                    ):

                        first_col_value = (
                            row[0]
                            if len(row) > 0
                            else None
                        )

                        if not _is_empty_cell(
                            first_col_value
                        ):
                            sheet_header_row = row_idx
                            break

            if sheet_header_row > max_row:

                workbook_data[sheet_name] = {
                    "headers": [],
                    "rows": [],
                    "detected_header_row": sheet_header_row,
                    "max_row": max_row,
                    "max_col": max_col,
                }

                continue

            header_idx = sheet_header_row - 1

            headers = values[header_idx]

            rows = values[
                header_idx + 1 :
            ]

            workbook_data[sheet_name] = {
                "headers": headers,
                "rows": rows,
                "detected_header_row": sheet_header_row,
                "max_row": max_row,
                "max_col": max_col,
            }

        return workbook_data


    # ==================================================
    # XLSX / XLSM
    # ==================================================
    wb = load_workbook(
        filename=file_obj,
        read_only=True,
        data_only=True,
    )

    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            workbook_data[ws.title] = {
                "headers": [],
                "rows": [],
                "detected_header_row": "",
                "max_row": max_row,
                "max_col": max_col,
            }
            continue

        sheet_header_row = (
            _detect_header_row(ws)
            if header_row is None
            else header_row
        )

        if sheet_header_row > max_row:
            workbook_data[ws.title] = {
                "headers": [],
                "rows": [],
                "detected_header_row": sheet_header_row,
                "max_row": max_row,
                "max_col": max_col,
            }
            continue

        headers = [
            ws.cell(row=sheet_header_row, column=col_idx).value
            for col_idx in range(1, max_col + 1)
        ]

        rows = []
        for row_idx in range(sheet_header_row + 1, max_row + 1):
            row_values = [
                ws.cell(row=row_idx, column=col_idx).value
                for col_idx in range(1, max_col + 1)
            ]
            rows.append(row_values)

        workbook_data[ws.title] = {
            "headers": headers,
            "rows": rows,
            "detected_header_row": sheet_header_row,
            "max_row": max_row,
            "max_col": max_col,
        }

    return workbook_data



# =========================================================
# Column handling
# =========================================================

def _build_columns(headers):
    """
    Build column metadata by header name/label.

    Important:
    - Compare columns by header value, not by position.
    - If duplicate labels exist, occurrence number is used internally.
    """
    columns = []
    occurrence_counter = defaultdict(int)

    for idx, header in enumerate(headers):
        header_token = _cell_token(header)
        occurrence_counter[header_token] += 1
        occurrence = occurrence_counter[header_token]

        display = _cell_display(header)
        if display == "":
            display = "BLANK_HEADER"

        col_id = f"{header_token}__occ{occurrence}"

        columns.append({
            "id": col_id,
            "header": header,
            "header_token": header_token,
            "display": display,
            "occurrence": occurrence,
            "position": idx,
        })

    return columns


def _make_report_column_names(common_columns):
    """
    Use original header display as report column names.
    If duplicated, add __2, __3...
    """
    name_count = defaultdict(int)
    report_names = []

    for col in common_columns:
        name = str(col["display"])
        name_count[name] += 1

        if name_count[name] == 1:
            report_names.append(name)
        else:
            report_names.append(f"{name}__{name_count[name]}")

    return report_names


def _compare_columns_by_name(sheet, columns_left, columns_right, labels):
    """
    Compare structure by variable/label name, not by column position.

    Columns only in one file are reported, but skipped for data comparison.
    """
    ids_left = {col["id"] for col in columns_left}
    ids_right = {col["id"] for col in columns_right}

    index_left_by_id = {
        col["id"]: col["position"]
        for col in columns_left
    }

    index_right_by_id = {
        col["id"]: col["position"]
        for col in columns_right
    }

    col_meta_left = {
        col["id"]: col
        for col in columns_left
    }

    col_meta_right = {
        col["id"]: col
        for col in columns_right
    }

    common_col_ids = [
        col["id"]
        for col in columns_left
        if col["id"] in ids_right
    ]

    only_left = [
        col["id"]
        for col in columns_left
        if col["id"] not in ids_right
    ]

    only_right = [
        col["id"]
        for col in columns_right
        if col["id"] not in ids_left
    ]

    structure_diff_rows = []

    for col_id in only_left:
        col = col_meta_left[col_id]
        structure_diff_rows.append({
            "SHEET": sheet,
            "DIFF_TYPE": f"Column only in {labels['LEFT']}",
            "COLUMN_NAME": col["display"],
            "OCCURRENCE": col["occurrence"],
            f"POSITION_{labels['LEFT'].upper()}": col["position"] + 1,
            f"POSITION_{labels['RIGHT'].upper()}": "",
        })

    for col_id in only_right:
        col = col_meta_right[col_id]
        structure_diff_rows.append({
            "SHEET": sheet,
            "DIFF_TYPE": f"Column only in {labels['RIGHT']}",
            "COLUMN_NAME": col["display"],
            "OCCURRENCE": col["occurrence"],
            f"POSITION_{labels['LEFT'].upper()}": "",
            f"POSITION_{labels['RIGHT'].upper()}": col["position"] + 1,
        })

    common_columns_left = [
        col_meta_left[col_id]
        for col_id in common_col_ids
    ]

    return (
        structure_diff_rows,
        common_columns_left,
        common_col_ids,
        index_left_by_id,
        index_right_by_id,
    )





def get_structure_diff_summary(
    structure_df,
    sheet_name,
    labels,
    summary_df=None,
):
    """
    Get structure difference summary for one sheet.
    """
    result = {
        "sheet_only": "",
        "left_only": "",
        "right_only": "",
    }


    # --------------------------------------------------
    # Sheet Difference
    # --------------------------------------------------
    if (
        summary_df is not None
        and not summary_df.empty
        and "SHEET" in summary_df.columns
        and "STATUS" in summary_df.columns
    ):

        summary_row = summary_df[
            summary_df["SHEET"] == sheet_name
        ]

        if not summary_row.empty:

            status = str(
                summary_row.iloc[0]["STATUS"]
            )

            if status.startswith("Only in"):
                result["sheet_only"] = status


    if structure_df.empty:
        return result

    required_cols = {
        "SHEET",
        "DIFF_TYPE",
        "COLUMN_NAME",
    }

    if not required_cols.issubset(
        set(structure_df.columns)
    ):
        return result

    sheet_df = structure_df[
        structure_df["SHEET"] == sheet_name
    ].copy()

    if sheet_df.empty:
        return result

    left = labels["LEFT"]
    right = labels["RIGHT"]

    left_cols = []
    right_cols = []

    for _, row in sheet_df.iterrows():

        diff_type = str(
            row.get("DIFF_TYPE", "")
        )

        column_name = str(
            row.get("COLUMN_NAME", "")
        )

        if not column_name.strip():
            continue

        diff_type_upper = diff_type.upper()

        # e.g.
        # "Column only in Main"
        # "Only in Main"
        if (
            "ONLY" in diff_type_upper
            and left.upper() in diff_type_upper
        ):
            left_cols.append(column_name)

        elif (
            "ONLY" in diff_type_upper
            and right.upper() in diff_type_upper
        ):
            right_cols.append(column_name)

    result["left_only"] = " | ".join(
        sorted(set(left_cols))
    )

    result["right_only"] = " | ".join(
        sorted(set(right_cols))
    )

    return result


# =========================================================
# Record comparison
# =========================================================

def _row_key_from_common_columns(row, common_col_ids, index_by_id):
    """
    Full-row key using all common columns.
    """
    tokens = []

    for col_id in common_col_ids:
        idx = index_by_id[col_id]
        value = row[idx] if idx < len(row) else None
        tokens.append(_cell_token(value))

    return tuple(tokens)


def _row_values_from_common_columns(row, common_col_ids, index_by_id):
    """
    Display values using common columns only.
    """
    values = []

    for col_id in common_col_ids:
        idx = index_by_id[col_id]
        value = row[idx] if idx < len(row) else None
        values.append(_cell_display(value))

    return values



def _build_record_counter(rows, common_col_ids, index_by_id, data_start_row):
    """
    Build multiset counter.

    Row order is ignored.
    Duplicate count is preserved.
    """
    counter = Counter()
    mapping = defaultdict(list)

    for zero_based_idx, row in enumerate(rows):
        excel_row_number = data_start_row + zero_based_idx

        key = _row_key_from_common_columns(
            row=row,
            common_col_ids=common_col_ids,
            index_by_id=index_by_id,
        )

        values = _row_values_from_common_columns(
            row=row,
            common_col_ids=common_col_ids,
            index_by_id=index_by_id,
        )

        counter[key] += 1

        mapping[key].append({
            "excel_row": excel_row_number,
            "values": values,
        })

    return counter, mapping




def _compare_records_by_full_common_columns(
    sheet,
    rows_left,
    rows_right,
    common_columns,
    common_col_ids,
    index_left_by_id,
    index_right_by_id,
    data_start_row_left,
    data_start_row_right,
    labels,
):
    """
    Compare data by full common-column row key.

    Data Diff output:
    - Duplicate Record
    - Only in Left
    - Only in Right
    - Record Count Difference

    """

    report_columns = _make_report_column_names(common_columns)

    counter_left, map_left = _build_record_counter(
        rows=rows_left,
        common_col_ids=common_col_ids,
        index_by_id=index_left_by_id,
        data_start_row=data_start_row_left,
    )

    counter_right, map_right = _build_record_counter(
        rows=rows_right,
        common_col_ids=common_col_ids,
        index_by_id=index_right_by_id,
        data_start_row=data_start_row_right,
    )

    diff_rows = []

    # --------------------------------------------------
    # Duplicate count for Summary
    # --------------------------------------------------
    duplicate_left_count = sum(
        count
        for count in counter_left.values()
        if count > 1
    )

    duplicate_right_count = sum(
        count
        for count in counter_right.values()
        if count > 1
    )

    # --------------------------------------------------
    # Duplicate keys
    # --------------------------------------------------
    duplicate_keys = {
        key
        for key in (
            set(counter_left.keys())
            | set(counter_right.keys())
        )
        if counter_left.get(key, 0) > 1
        or counter_right.get(key, 0) > 1
    }


    # --------------------------------------------------
    # Output Duplicate Records
    # Every duplicate row is listed.
    # --------------------------------------------------
    for key, count in counter_left.items():

        if count > 1:

            for row_info in map_left[key]:

                record = {
                    "DIFF_TYPE": "Duplicate Record",
                    "SOURCE": labels["LEFT"],
                }

                for col_name, value in zip(
                    report_columns,
                    row_info["values"],
                ):
                    record[col_name] = value

                diff_rows.append(record)

    for key, count in counter_right.items():

        if count > 1:

            for row_info in map_right[key]:

                record = {
                    "DIFF_TYPE": "Duplicate Record",
                    "SOURCE": labels["RIGHT"],
                }

                for col_name, value in zip(
                    report_columns,
                    row_info["values"],
                ):
                    record[col_name] = value

                diff_rows.append(record)

    # --------------------------------------------------
    # Compare all row keys
    # --------------------------------------------------
    all_keys = sorted(
        set(counter_left.keys()) | set(counter_right.keys()),
        key=lambda x: repr(x),
    )

    only_left_count = 0
    only_right_count = 0

    for key in all_keys:
        
        if key in duplicate_keys:
            continue

        count_left = counter_left.get(key, 0)
        count_right = counter_right.get(key, 0)

        # completely identical
        if count_left == count_right:
            continue

        rows_info_left = map_left.get(key, [])
        rows_info_right = map_right.get(key, [])

        # --------------------------------------------------
        # Only in LEFT
        # --------------------------------------------------
        if count_left > 0 and count_right == 0:

            only_left_count += count_left

            for row_info in rows_info_left:

                record = {
                    "DIFF_TYPE": labels["ONLY_LEFT"],
                    "SOURCE": labels["LEFT"],
                }

                for col_name, value in zip(
                    report_columns,
                    row_info["values"],
                ):
                    record[col_name] = value

                diff_rows.append(record)

        # --------------------------------------------------
        # Only in RIGHT
        # --------------------------------------------------
        elif count_left == 0 and count_right > 0:

            only_right_count += count_right

            for row_info in rows_info_right:

                record = {
                    "DIFF_TYPE": labels["ONLY_RIGHT"],
                    "SOURCE": labels["RIGHT"],
                }

                for col_name, value in zip(
                    report_columns,
                    row_info["values"],
                ):
                    record[col_name] = value

                diff_rows.append(record)

        # --------------------------------------------------
        # Same key but different count
        # --------------------------------------------------
        else:

            max_count = max(
                len(rows_info_left),
                len(rows_info_right),
            )

            for i in range(max_count):

                if i < len(rows_info_left):

                    record = {
                        "DIFF_TYPE": "Record Count Difference",
                        "SOURCE": labels["LEFT"],
                    }

                    for col_name, value in zip(
                        report_columns,
                        rows_info_left[i]["values"],
                    ):
                        record[col_name] = value

                    diff_rows.append(record)

                if i < len(rows_info_right):

                    record = {
                        "DIFF_TYPE": "Record Count Difference",
                        "SOURCE": labels["RIGHT"],
                    }

                    for col_name, value in zip(
                        report_columns,
                        rows_info_right[i]["values"],
                    ):
                        record[col_name] = value

                    diff_rows.append(record)

    stats = {
        "N_RECORD_DIFF_ROWS": len(diff_rows),

        f"N_ONLY_IN_{labels['LEFT'].upper()}_RECORDS":
            only_left_count,

        f"N_ONLY_IN_{labels['RIGHT'].upper()}_RECORDS":
            only_right_count,

        f"N_DUP_ROW_{labels['LEFT'].upper()}":
            duplicate_left_count,

        f"N_DUP_ROW_{labels['RIGHT'].upper()}":
            duplicate_right_count,
    }

    return pd.DataFrame(diff_rows), stats




# =========================================================
# Main function
# =========================================================

def compare_excel_verify(
    file_left,
    file_right,
    compare_type="Main/Double Comparison",
    header_row=None,
):
    """
    Main Verify function.

    Verify spec:
    1. Sheet diff is summarized in Summary.
    2. Sheets only in one file are skipped.
    3. Columns are compared by variable/label name, not by position.
    4. Columns only in one file are skipped for data comparison.
    5. Data comparison uses all common columns as the full-row key.
    6. Output report is organized by sheet.
    7. If header_row is None, header row is auto-detected per sheet.
    """

    labels = _get_compare_labels(compare_type)

    wb_left = _read_excel(file_left, header_row=header_row)
    wb_right = _read_excel(file_right, header_row=header_row)

    sheets_left = set(wb_left.keys())
    sheets_right = set(wb_right.keys())

    only_sheets_left = sorted(sheets_left - sheets_right)
    only_sheets_right = sorted(sheets_right - sheets_left)
    common_sheets = sorted(sheets_left & sheets_right)

    summary_rows = []
    structure_diff_rows = []
    sheet_reports = {}

    # =====================================================
    # Sheets only in LEFT
    # =====================================================
    for sheet in only_sheets_left:
        table_left = wb_left[sheet]

        summary_rows.append({
            "SHEET": sheet,
            "STATUS": f"Only in {labels['LEFT']}",
            "NOTE": f"Sheet exists only in {labels['LEFT']}. Data comparison skipped.",

            f"HEADER_ROW_{labels['LEFT'].upper()}": table_left.get("detected_header_row", ""),
            f"HEADER_ROW_{labels['RIGHT'].upper()}": "",

            f"N_ROWS_{labels['LEFT'].upper()}": len(table_left["rows"]),
            f"N_ROWS_{labels['RIGHT'].upper()}": "",

            f"N_COLS_{labels['LEFT'].upper()}": len(table_left["headers"]),
            f"N_COLS_{labels['RIGHT'].upper()}": "",

            "N_COMMON_COLUMNS": "",
            "N_STRUCTURE_DIFF": "",
            "N_RECORD_DIFF_ROWS": "",

            f"N_ONLY_IN_{labels['LEFT'].upper()}_RECORDS": "",
            f"N_ONLY_IN_{labels['RIGHT'].upper()}_RECORDS": "",

            f"N_DUP_ROW_{labels['LEFT'].upper()}": "",
            f"N_DUP_ROW_{labels['RIGHT'].upper()}": "",
        })

    # =====================================================
    # Sheets only in RIGHT
    # =====================================================
    for sheet in only_sheets_right:
        table_right = wb_right[sheet]

        summary_rows.append({
            "SHEET": sheet,
            "STATUS": f"Only in {labels['RIGHT']}",
            "NOTE": f"Sheet exists only in {labels['RIGHT']}. Data comparison skipped.",

            f"HEADER_ROW_{labels['LEFT'].upper()}": "",
            f"HEADER_ROW_{labels['RIGHT'].upper()}": table_right.get("detected_header_row", ""),

            f"N_ROWS_{labels['LEFT'].upper()}": "",
            f"N_ROWS_{labels['RIGHT'].upper()}": len(table_right["rows"]),

            f"N_COLS_{labels['LEFT'].upper()}": "",
            f"N_COLS_{labels['RIGHT'].upper()}": len(table_right["headers"]),

            "N_COMMON_COLUMNS": "",
            "N_STRUCTURE_DIFF": "",
            "N_RECORD_DIFF_ROWS": "",

            f"N_ONLY_IN_{labels['LEFT'].upper()}_RECORDS": "",
            f"N_ONLY_IN_{labels['RIGHT'].upper()}_RECORDS": "",

            f"N_DUP_ROW_{labels['LEFT'].upper()}": "",
            f"N_DUP_ROW_{labels['RIGHT'].upper()}": "",
        })

    # =====================================================
    # Compare common sheets
    # =====================================================
    for sheet in common_sheets:
        table_left = wb_left[sheet]
        table_right = wb_right[sheet]

        headers_left = table_left["headers"]
        headers_right = table_right["headers"]

        rows_left = table_left["rows"]
        rows_right = table_right["rows"]

        header_row_left = table_left.get("detected_header_row", "")
        header_row_right = table_right.get("detected_header_row", "")

        data_start_row_left = "" if header_row_left == "" else header_row_left + 1
        data_start_row_right = "" if header_row_right == "" else header_row_right + 1

        columns_left = _build_columns(headers_left)
        columns_right = _build_columns(headers_right)

        (
            sheet_structure_diff,
            common_columns,
            common_col_ids,
            index_left_by_id,
            index_right_by_id,
        ) = _compare_columns_by_name(
            sheet=sheet,
            columns_left=columns_left,
            columns_right=columns_right,
            labels=labels,
        )

        structure_diff_rows.extend(sheet_structure_diff)

        # =================================================
        # Data compare
        # =================================================
        if not common_col_ids:
            diff_df = pd.DataFrame()
            record_stats = {
                "N_RECORD_DIFF_ROWS": "",
                f"N_ONLY_IN_{labels['LEFT'].upper()}_RECORDS": "",
                f"N_ONLY_IN_{labels['RIGHT'].upper()}_RECORDS": "",
                f"N_DUP_ROW_{labels['LEFT'].upper()}": "",
                f"N_DUP_ROW_{labels['RIGHT'].upper()}": "",
            }
            note = "No common columns found. Data comparison skipped."

        elif data_start_row_left == "" or data_start_row_right == "":
            diff_df = pd.DataFrame()
            record_stats = {
                "N_RECORD_DIFF_ROWS": "",
                f"N_ONLY_IN_{labels['LEFT'].upper()}_RECORDS": "",
                f"N_ONLY_IN_{labels['RIGHT'].upper()}_RECORDS": "",
                f"N_DUP_ROW_{labels['LEFT'].upper()}": "",
                f"N_DUP_ROW_{labels['RIGHT'].upper()}": "",
            }
            note = "Header row could not be detected. Data comparison skipped."

        else:
            diff_df, record_stats = _compare_records_by_full_common_columns(
                sheet=sheet,
                rows_left=rows_left,
                rows_right=rows_right,
                common_columns=common_columns,
                common_col_ids=common_col_ids,
                index_left_by_id=index_left_by_id,
                index_right_by_id=index_right_by_id,
                data_start_row_left=data_start_row_left,
                data_start_row_right=data_start_row_right,
                labels=labels,
            )
            note = ""

        
        if diff_df.empty:

            report_columns = _make_report_column_names(
                common_columns
            )

            diff_df = pd.DataFrame(
                columns=[
                    "DIFF_TYPE",
                    "SOURCE",
                    *report_columns,
                ]
            )

        sheet_reports[sheet] = diff_df

        has_structure_diff = len(sheet_structure_diff) > 0

        has_record_diff = (
            record_stats["N_RECORD_DIFF_ROWS"] != ""
            and record_stats["N_RECORD_DIFF_ROWS"] > 0
        )

        has_dup_row = (
            record_stats[f"N_DUP_ROW_{labels['LEFT'].upper()}"] != ""
            and record_stats[f"N_DUP_ROW_{labels['LEFT'].upper()}"] > 0
        ) or (
            record_stats[f"N_DUP_ROW_{labels['RIGHT'].upper()}"] != ""
            and record_stats[f"N_DUP_ROW_{labels['RIGHT'].upper()}"] > 0
        )

        has_data_diff = has_record_diff or has_dup_row

        if has_structure_diff and has_data_diff:
            status = "Inconsistent (Structure&Data)"
        elif has_structure_diff:
            status = "Inconsistent (Structure)"
        elif has_data_diff:
            status = "Inconsistent (Data)"
        else:
            status = "Consistent"

        if has_structure_diff:
            add_note = "Columns differ between both files. Only common columns were compared."
            note = f"{note} {add_note}".strip()

        summary_rows.append({
            "SHEET": sheet,
            "STATUS": status,
            "NOTE": note,

            f"HEADER_ROW_{labels['LEFT'].upper()}": header_row_left,
            f"HEADER_ROW_{labels['RIGHT'].upper()}": header_row_right,

            f"N_ROWS_{labels['LEFT'].upper()}": len(rows_left),
            f"N_ROWS_{labels['RIGHT'].upper()}": len(rows_right),

            f"N_COLS_{labels['LEFT'].upper()}": len(headers_left),
            f"N_COLS_{labels['RIGHT'].upper()}": len(headers_right),

            "N_COMMON_COLUMNS": len(common_col_ids),
            "N_STRUCTURE_DIFF": len(sheet_structure_diff),

            "N_RECORD_DIFF_ROWS": record_stats["N_RECORD_DIFF_ROWS"],

            f"N_ONLY_IN_{labels['LEFT'].upper()}_RECORDS": record_stats[
                f"N_ONLY_IN_{labels['LEFT'].upper()}_RECORDS"
            ],
            f"N_ONLY_IN_{labels['RIGHT'].upper()}_RECORDS": record_stats[
                f"N_ONLY_IN_{labels['RIGHT'].upper()}_RECORDS"
            ],

            f"N_DUP_ROW_{labels['LEFT'].upper()}": record_stats[
                f"N_DUP_ROW_{labels['LEFT'].upper()}"
            ],
            f"N_DUP_ROW_{labels['RIGHT'].upper()}": record_stats[
                f"N_DUP_ROW_{labels['RIGHT'].upper()}"
            ],
        })

    result = {
        "Compare_Type": compare_type,
        "Labels": labels,
        "Summary": pd.DataFrame(summary_rows),
        "Structure_Diff": pd.DataFrame(structure_diff_rows),
        "Sheet_Reports": sheet_reports,
    }

    return result



# =========================================================
# Summary
# =========================================================
def build_summary_view(result):
    summary = result.get("Summary", pd.DataFrame()).copy()
    structure = result.get("Structure_Diff", pd.DataFrame()).copy()

    if summary.empty:
        return summary

    labels = result.get("Labels", {})
    left = labels.get("LEFT", "Main")
    right = labels.get("RIGHT", "Double")

    left_u = left.upper()
    right_u = right.upper()

    summary = summary.sort_values(by="SHEET").reset_index(drop=True)

    # --------------------------------------------------
    # Diff Variables 分 A/B，只抓 Column 差異
    # --------------------------------------------------
    if (
        not structure.empty
        and "COLUMN_NAME" in structure.columns
        and "DIFF_TYPE" in structure.columns
    ):

        diff_left_df = structure[
            structure["DIFF_TYPE"].astype(str).str.contains(
                "Column",
                case=False,
                na=False,
            )
            & structure["DIFF_TYPE"].astype(str).str.contains(
                left,
                case=False,
                na=False,
            )
        ]

        diff_right_df = structure[
            structure["DIFF_TYPE"].astype(str).str.contains(
                "Column",
                case=False,
                na=False,
            )
            & structure["DIFF_TYPE"].astype(str).str.contains(
                right,
                case=False,
                na=False,
            )
        ]

        diff_map_left = (
            diff_left_df.groupby("SHEET")["COLUMN_NAME"]
            .apply(lambda x: " | ".join(sorted(set(map(str, x)))))
            .to_dict()
        )

        diff_map_right = (
            diff_right_df.groupby("SHEET")["COLUMN_NAME"]
            .apply(lambda x: " | ".join(sorted(set(map(str, x)))))
            .to_dict()
        )

        summary[f"Diff Variables ({left})"] = (
            summary["SHEET"].map(diff_map_left).fillna("")
        )
        summary[f"Diff Variables ({right})"] = (
            summary["SHEET"].map(diff_map_right).fillna("")
        )

    else:
        summary[f"Diff Variables ({left})"] = ""
        summary[f"Diff Variables ({right})"] = ""

    rename_map = {}

    if f"N_COLS_{left_u}" in summary.columns:
        rename_map[f"N_COLS_{left_u}"] = f"Number of Variable ({left})"

    if f"N_COLS_{right_u}" in summary.columns:
        rename_map[f"N_COLS_{right_u}"] = f"Number of Variable ({right})"

    if f"N_ROWS_{left_u}" in summary.columns:
        rename_map[f"N_ROWS_{left_u}"] = f"Number of Row ({left})"

    if f"N_ROWS_{right_u}" in summary.columns:
        rename_map[f"N_ROWS_{right_u}"] = f"Number of Row ({right})"

    if f"N_DUP_ROW_{left_u}" in summary.columns:
        rename_map[f"N_DUP_ROW_{left_u}"] = f"Number of Dup. Row ({left})"

    if f"N_DUP_ROW_{right_u}" in summary.columns:
        rename_map[f"N_DUP_ROW_{right_u}"] = f"Number of Dup. Row ({right})"

    if "N_RECORD_DIFF_ROWS" in summary.columns:
        rename_map["N_RECORD_DIFF_ROWS"] = "Number of Different Rows"

    summary = summary.rename(columns=rename_map)

    numeric_cols = [c for c in summary.columns if "Number" in c]

    for col in numeric_cols:
        summary[col] = (
            pd.to_numeric(summary[col], errors="coerce")
            .fillna(0)
            .astype(int)
        )

    summary = summary.rename(columns={
        "SHEET": "Sheet",
        "STATUS": "Status",
        "NOTE": "Note",
    })

    display_cols = [
        "Sheet",
        "Status",

        f"Number of Variable ({left})",
        f"Number of Variable ({right})",

        f"Number of Row ({left})",
        f"Number of Row ({right})",

        f"Number of Dup. Row ({left})",
        f"Number of Dup. Row ({right})",

        f"Diff Variables ({left})",
        f"Diff Variables ({right})",

        "Number of Different Rows",

        "Note",
    ]

    display_cols = [c for c in display_cols if c in summary.columns]

    return summary[display_cols]



#Write Summary sheet with merged multi-level headers.
def _write_merged_summary_sheet(writer, result, sheet_name="Summary"):

    summary_view = build_summary_view(result)

    labels = result.get("Labels", {})
    left = labels.get("LEFT", "Main")
    right = labels.get("RIGHT", "Double")

    workbook = writer.book
    ws = workbook.create_sheet(title=sheet_name)

    # --------------------------------------------------
    # Styles
    # --------------------------------------------------
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_fill = PatternFill("solid", fgColor="EAF3F8")
    gray_fill = PatternFill("solid", fgColor="F2F2F2")

    red_font = Font(name="Calibri", size=10, color="FF0000")
    black_font = Font(name="Calibri", size=10, color="000000")
    header_font = Font(name="Calibri", size=10, bold=True, color="000000")

    center_align = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    left_align = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --------------------------------------------------
    # Header rows
    # --------------------------------------------------
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:D1")
    ws.merge_cells("E1:F1")
    ws.merge_cells("G1:H1")
    ws.merge_cells("I1:J1")
    ws.merge_cells("K1:K2")
    ws.merge_cells("L1:L2")

    ws["A1"] = "Sheet"
    ws["B1"] = "Status"

    ws["C1"] = "Number of Variable"
    ws["C2"] = left
    ws["D2"] = right

    ws["E1"] = "Number of Row"
    ws["E2"] = left
    ws["F2"] = right

    ws["G1"] = "Number of Dup. Row"
    ws["G2"] = left
    ws["H2"] = right

    ws["I1"] = "Diff Variables"
    ws["I2"] = left
    ws["J2"] = right

    ws["K1"] = "Number of Different Rows"
    ws["L1"] = "Note"

    # only style writable cells to avoid MergedCell error
    header_cells = [
        "A1", "B1", "C1", "E1", "G1", "I1", "K1", "L1",
        "C2", "D2", "E2", "F2", "G2", "H2", "I2", "J2",
    ]

    for cell_ref in header_cells:
        cell = ws[cell_ref]
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        cell.fill = header_fill if cell.row == 1 else subheader_fill

    # --------------------------------------------------
    # Data columns
    # --------------------------------------------------
    export_columns = [
        "Sheet",
        "Status",
        f"Number of Variable ({left})",
        f"Number of Variable ({right})",
        f"Number of Row ({left})",
        f"Number of Row ({right})",
        f"Number of Dup. Row ({left})",
        f"Number of Dup. Row ({right})",
        f"Diff Variables ({left})",
        f"Diff Variables ({right})",
        "Number of Different Rows",
        "Note",
    ]

    export_columns = [
        c for c in export_columns
        if c in summary_view.columns
    ]

    col_index = {
        col_name: idx + 1
        for idx, col_name in enumerate(export_columns)
    }

    def _to_number(value):
        try:
            if value in ["", None]:
                return 0
            return float(value)
        except Exception:
            return 0

    def _is_non_empty(value):
        return value is not None and str(value).strip() != ""

    def _mark_red(excel_row, col_name):
        if col_name in col_index:
            ws.cell(row=excel_row, column=col_index[col_name]).font = red_font

    def _apply_gray_row(excel_row):
        for col_idx in range(1, len(export_columns) + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = gray_fill

    start_row = 3

    for row_offset, (_, row) in enumerate(summary_view.iterrows()):
        excel_row = start_row + row_offset
        status = str(row.get("Status", ""))

        # --------------------------------------------------
        # Write values
        # --------------------------------------------------
        for col_idx, col_name in enumerate(export_columns, start=1):
            value = row.get(col_name, "")

            if isinstance(value, float):
                value = int(value)

            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = value
            cell.alignment = left_align
            cell.border = border
            cell.font = black_font

        # --------------------------------------------------
        # Whole inconsistent row = gray background only
        # --------------------------------------------------
        if status != "Consistent":
            _apply_gray_row(excel_row)

        # --------------------------------------------------
        # Red font only for true inconsistent cells
        # --------------------------------------------------

        # Number of Variable pair
        variable_left = f"Number of Variable ({left})"
        variable_right = f"Number of Variable ({right})"

        if variable_left in export_columns and variable_right in export_columns:
            if _to_number(row.get(variable_left, 0)) != _to_number(row.get(variable_right, 0)):
                _mark_red(excel_row, variable_left)
                _mark_red(excel_row, variable_right)

        # Number of Row pair
        row_left = f"Number of Row ({left})"
        row_right = f"Number of Row ({right})"

        if row_left in export_columns and row_right in export_columns:
            if _to_number(row.get(row_left, 0)) != _to_number(row.get(row_right, 0)):
                _mark_red(excel_row, row_left)
                _mark_red(excel_row, row_right)

        # Number of Dup. Row side-specific
        dup_left = f"Number of Dup. Row ({left})"
        dup_right = f"Number of Dup. Row ({right})"

        if dup_left in export_columns and _to_number(row.get(dup_left, 0)) > 0:
            _mark_red(excel_row, dup_left)

        if dup_right in export_columns and _to_number(row.get(dup_right, 0)) > 0:
            _mark_red(excel_row, dup_right)

        # Diff Variables side-specific
        diff_var_left = f"Diff Variables ({left})"
        diff_var_right = f"Diff Variables ({right})"

        if diff_var_left in export_columns and _is_non_empty(row.get(diff_var_left, "")):
            _mark_red(excel_row, diff_var_left)

        if diff_var_right in export_columns and _is_non_empty(row.get(diff_var_right, "")):
            _mark_red(excel_row, diff_var_right)

        # Number of Different Rows
        diff_rows_col = "Number of Different Rows"

        if diff_rows_col in export_columns and _to_number(row.get(diff_rows_col, 0)) > 0:
            _mark_red(excel_row, diff_rows_col)

        # NOTE stays black
        if "Note" in export_columns:
            ws.cell(row=excel_row, column=col_index["Note"]).font = black_font

    # --------------------------------------------------
    # Freeze / filter / width
    # --------------------------------------------------
    ws.freeze_panes = "A3"

    if ws.max_row >= 3:
        ws.auto_filter.ref = f"A2:L{ws.max_row}"

    width_map = {
        "A": 18,
        "B": 28,
        "C": 15,
        "D": 15,
        "E": 15,
        "F": 15,
        "G": 15,
        "H": 15,
        "I": 45,
        "J": 45,
        "K": 15,
        "L": 55,
    }

    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24





def _auto_format_regular_sheet(ws, red_all_data_rows=False):
    """
    Apply basic formatting to regular sheets.
    """

    if ws.max_row < 1 or ws.max_column < 1:
        return

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Calibri", size=10, bold=True, color="000000")
    red_font = Font(name="Calibri", size=10, color="FF0000")


    center_align = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    left_align = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = left_align
            cell.border = border

            if red_all_data_rows:
                cell.font = red_font

    ws.freeze_panes = "A2"

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value

            if value is None:
                continue

            max_len = max(max_len, len(str(value)))

        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, 12),
            60,
        )



def _apply_data_diff_styling(ws, labels):
    """
    Apply styling to Data Diff sheets.
    """


    if ws.max_row < 2 or ws.max_column < 1:
        return

    left = labels.get("LEFT", "Main")
    right = labels.get("RIGHT", "Double")

    left_u = left.upper()
    right_u = right.upper()

    only_left_label = labels.get("ONLY_LEFT", f"Only in {left}")
    only_right_label = labels.get("ONLY_RIGHT", f"Only in {right}")

    # --------------------------------------------------
    # Styles
    # --------------------------------------------------
    gray_fill = PatternFill(
        "solid",
        fgColor="F2F2F2",
    )

    green_fill = PatternFill(
        "solid",
        fgColor="E2F0D9",
    )

    orange_fill = PatternFill(
        "solid",
        fgColor="FCE4D6",
    )

    red_font = Font(
        name="Calibri",
        size=10,
        color="FF0000",
    )

    black_font = Font(
        name="Calibri",
        size=10,
        color="000000",
    )

    # --------------------------------------------------
    # Header map
    # --------------------------------------------------
    headers = [
        ws.cell(row=1, column=col_idx).value
        for col_idx in range(1, ws.max_column + 1)
    ]

    header_idx = {
        str(header): idx + 1
        for idx, header in enumerate(headers)
        if header is not None
    }

    diff_type_col = header_idx.get("DIFF_TYPE")
    source_col = header_idx.get("SOURCE")

    count_left_col_name = f"COUNT_{left_u}"
    count_right_col_name = f"COUNT_{right_u}"

    rows_left_col_name = f"ROWS_{left_u}"
    rows_right_col_name = f"ROWS_{right_u}"

    count_left_col = header_idx.get(count_left_col_name)
    count_right_col = header_idx.get(count_right_col_name)

    rows_left_col = header_idx.get(rows_left_col_name)
    rows_right_col = header_idx.get(rows_right_col_name)

    metadata_cols = {
        "DIFF_TYPE",
        "SOURCE",
        count_left_col_name,
        count_right_col_name,
        rows_left_col_name,
        rows_right_col_name,
    }

    value_col_indices = [
        col_idx
        for col_idx, header in enumerate(headers, start=1)
        if str(header) not in metadata_cols
    ]

    # --------------------------------------------------
    # Helpers
    # --------------------------------------------------
    def _cell_text(row_idx, col_idx):
        value = ws.cell(row=row_idx, column=col_idx).value
        return "" if value is None else str(value)

    def _get_diff_type(row_idx):
        if not diff_type_col:
            return ""
        return _cell_text(row_idx, diff_type_col)

    def _get_source(row_idx):
        if not source_col:
            return ""
        return _cell_text(row_idx, source_col)

    def _set_black(row_idx, col_idx):
        if col_idx:
            ws.cell(row=row_idx, column=col_idx).font = black_font

    def _set_red(row_idx, col_idx):
        if col_idx:
            ws.cell(row=row_idx, column=col_idx).font = red_font

    def _gray_row(row_idx):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = gray_fill

    def _green_row(row_idx):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = green_fill

    def _orange_row(row_idx):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = orange_fill

    def _row_value_map(row_idx):
        return {
            col_idx: _cell_text(row_idx, col_idx)
            for col_idx in value_col_indices
        }

    def _similarity_score(row_a, row_b):
        """
        Pair rows by highest number of identical original value cells.
        This is used to pair:
            Only in Main/New
            Only in Double/Previous
        """
        values_a = _row_value_map(row_a)
        values_b = _row_value_map(row_b)

        score = 0

        for col_idx in value_col_indices:
            if values_a.get(col_idx, "") == values_b.get(col_idx, ""):
                score += 1

        return score

    # --------------------------------------------------
    # Initial styling and collect row groups
    # --------------------------------------------------
    left_rows = []
    right_rows = []
    record_count_diff_rows = []

    for row_idx in range(2, ws.max_row + 1):

        # Default all cells to Calibri 10 black
        for col_idx in range(1, ws.max_column + 1):
            _set_black(row_idx, col_idx)

        source = _get_source(row_idx)
        diff_type = _get_diff_type(row_idx)

        if diff_type == "Duplicate Record":
            if source == left:
                _green_row(row_idx)

            elif source == right:
                _orange_row(row_idx)

        elif source == right:
            _gray_row(row_idx)

        if diff_type == only_left_label:
            left_rows.append(row_idx)

        elif diff_type == only_right_label:
            right_rows.append(row_idx)

        elif diff_type == "Record Count Difference":
            record_count_diff_rows.append(row_idx)

    # --------------------------------------------------
    # Pair Only-in-left and Only-in-right rows
    # Then mark only truly different original value cells red
    # --------------------------------------------------
    used_right_rows = set()

    for left_row in left_rows:

        best_right_row = None
        best_score = -1

        for right_row in right_rows:
            if right_row in used_right_rows:
                continue

            score = _similarity_score(left_row, right_row)

            if score > best_score:
                best_score = score
                best_right_row = right_row

        # No paired right row: entire original value cells are unmatched
        if best_right_row is None:
            for col_idx in value_col_indices:
                _set_red(left_row, col_idx)
            continue

        used_right_rows.add(best_right_row)

        # Compare paired rows cell by cell
        for col_idx in value_col_indices:
            left_value = _cell_text(left_row, col_idx)
            right_value = _cell_text(best_right_row, col_idx)

            if left_value != right_value:
                _set_red(left_row, col_idx)
                _set_red(best_right_row, col_idx)

    # Right rows without paired left row:
    # entire original value cells are unmatched
    for right_row in right_rows:
        if right_row not in used_right_rows:
            for col_idx in value_col_indices:
                _set_red(right_row, col_idx)

    # --------------------------------------------------
    # Record Count Difference
    # Data values are the same key; only count/rows differ.
    # Mark COUNT / ROWS red only.
    # --------------------------------------------------
    for row_idx in record_count_diff_rows:
        _set_red(row_idx, count_left_col)
        _set_red(row_idx, count_right_col)
        _set_red(row_idx, rows_left_col)
        _set_red(row_idx, rows_right_col)

    # --------------------------------------------------
    # Keep metadata cells black unless Record Count Difference changed them.
    # For Only in rows, DIFF_TYPE / SOURCE / COUNT / ROWS stay black.
    # --------------------------------------------------
    for row_idx in list(left_rows) + list(right_rows):
        _set_black(row_idx, diff_type_col)
        _set_black(row_idx, source_col)
        _set_black(row_idx, count_left_col)
        _set_black(row_idx, count_right_col)
        _set_black(row_idx, rows_left_col)
        _set_black(row_idx, rows_right_col)




# =========================================================
# Export
# =========================================================
def export_verify_report(result):

    output = io.BytesIO()
    used_sheet_names = set()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # --------------------------------------------------
        # Summary with merged header
        # --------------------------------------------------
        summary_sheet_name = _safe_sheet_name("Summary", used_sheet_names)

        _write_merged_summary_sheet(
            writer=writer,
            result=result,
            sheet_name=summary_sheet_name,
        )

        workbook = writer.book
        

        # --------------------------------------------------
        # Sheet-level Data Diff sheets
        # --------------------------------------------------
        sheet_reports = result.get("Sheet_Reports", {})

        for source_sheet_name, df in sheet_reports.items():

            report_sheet_name = _safe_sheet_name(
                source_sheet_name,
                used_sheet_names,
            )
            
            # -----------------------------
            # Write data diff
            # -----------------------------
            df.to_excel(
                writer,
                sheet_name=report_sheet_name,
                index=False,
            )

            ws = workbook[report_sheet_name]

            
            summary_df = result.get(
                "Summary",
                pd.DataFrame(),
            )

            status = "Consistent"

            sheet_row = summary_df[
                summary_df["SHEET"] == source_sheet_name
            ]

            if (
                not sheet_row.empty
                and "STATUS" in sheet_row.columns
            ):
                status = str(
                    sheet_row.iloc[0]["STATUS"]
                )

            if status != "Consistent":
                # Red tab
                ws.sheet_properties.tabColor = "FF0000"


            # -----------------------------
            # Excel formatting
            # -----------------------------
            _auto_format_regular_sheet(
                ws,
                red_all_data_rows=False,
            )

            _apply_data_diff_styling(
                ws=ws,
                labels=result["Labels"],
            )

            # -----------------------------
            # Structure Difference
            # append to bottom
            # -----------------------------
            structure_info = get_structure_diff_summary(
                structure_df=result.get(
                    "Structure_Diff",
                    pd.DataFrame(),
                ),
                sheet_name=source_sheet_name,
                labels=result["Labels"],
                summary_df=result.get(
                    "Summary",
                    pd.DataFrame(),
                ),
            )

            info_rows = []

            if structure_info["sheet_only"]:
                info_rows.append(
                    (
                        "Sheet Difference",
                        structure_info["sheet_only"],
                    )
                )

            if (
                structure_info["left_only"]
                or structure_info["right_only"]
            ):
                info_rows.append(
                    (
                        f"{result['Labels']['LEFT']} Only",
                        structure_info["left_only"],
                    )
                )

                info_rows.append(
                    (
                        f"{result['Labels']['RIGHT']} Only",
                        structure_info["right_only"],
                    )
                )

            if info_rows:

                start_row = ws.max_row + 3

                title_cell = ws.cell(
                    row=start_row,
                    column=1,
                )

                title_cell.value = "Structure Difference"

                title_cell.font = Font(
                    name="Calibri",
                    size=10,
                    bold=True,
                )

                title_cell.fill = PatternFill(
                    "solid",
                    fgColor="D9EAF7",
                )

                for offset, (label, value) in enumerate(
                    info_rows,
                    start=1,
                ):

                    ws.cell(
                        row=start_row + offset,
                        column=1,
                    ).value = label

                    ws.cell(
                        row=start_row + offset,
                        column=2,
                    ).value = value

                    ws.cell(
                        row=start_row + offset,
                        column=1,
                    ).font = Font(
                        name="Calibri",
                        size=10,
                    )

                    ws.cell(
                        row=start_row + offset,
                        column=2,
                    ).font = Font(
                        name="Calibri",
                        size=10,
                    )

    output.seek(0)

    return output





